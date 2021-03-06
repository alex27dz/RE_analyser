import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
import mysql.connector
import datetime
import json  # working with json dicts
import yagmail  # importing all email file to use send function
import datetime  # datetime.datetime.now()
import time
import pprint
import logging
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl.styles import Alignment
import urllib.request
import uuid
from selenium.webdriver.common.action_chains import ActionChains


class Builders(object):
    def __init__(self, metropolitan, short_state, xls_name):
        self.driver = webdriver.Chrome("/Users/alexdezho/Downloads/chromedriver")
        self.lennar_url = 'https://www.lennar.com/'  # builders website
        self.metropolitan = metropolitan.lower() + ' ' + short_state.lower()  # full name for search
        self.floorplan_homes = ''
        self.xls_name = xls_name  # xls name
        self.short_state = short_state
        self.clicked = ''
        self.list_of_homes = []
        self.community_address_list_full = []  # full list of community addresses
        self.community_address_list_names = []  # full list of community names
        self.id_random_list = []
        self.row = 2
        self.rowhome = 2
        self.general_row = 0
        self.row_num_xls = 0
        self.index = 1
        self.x_path_name_to_scroll = ''
        self.element = ''  # scrolling element
        self.addr = ''
        self.name = ''
        self.update_time = ''  # update time
        self.homes_urls = []  # list of homes urls
        self.x_path_name = ''
        self.num_of_communities = ''
        self.num_of_pages = ''  # num of community pages
        self.num_of_comm_pages = ''
        self.num_of_homes_pages = ''  # num of homes pages
        self.num_of_moving_homes = ''  # num of homes
        self.dict_lennar_filter_info = {
            'Communities num': '',
            'metropolitan name': metropolitan,
            'Quick Move-In Homes num': '',
            'Floorplans num': '',
            'time of update': ''
        }

        # community data for mysql and xls
        self.dict_community_data = {
            'address': '',
            'name_community': '',
            'overview': '',
            'approximate_hoa_fees': 'non',
            'approximate_tax_rate': 'non',
            'included_features_pdf_url': 'under solution',
            'community_map_url': 'no pic',
            'community_home_picture_for_present_url': 'no pic',
            'available_homes_quick_move_in_homes': '',
            'available_homes_floorplans': '',
            'id_generated':''

        }

        # home data for mysql and xls
        self.dict_home_data = {
            'address': '',
            'name_community': metropolitan,
            'home_name': '',
            'home_site': '',
            'availability': '',
            'priced_from': '',
            'home_size': '',
            'stories': '',
            'beds': '',
            'type': '',
            'baths': '',
            'garage': '',
            'id': '',
            'id_generated': '',
            'description': '',
            'included_features_pdf': 'under solution',
            'floorplans_with_furniture_pic': '',
            'id_generated_home': '',
            'gallery_view_picture': ''
        }

    def closeBrowser(self):
        self.driver.close()
        logging.debug('Browser closed')
        print('Browser closed')

    def lennar_filter_and_toolbar_info_copy(self):
        driver = self.driver
        driver.get(self.lennar_url)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[1]/div/div/div[2]/aside/div/input')))  # await command
        driver.find_element_by_xpath('//*[@id="wrapper"]/section[1]/div/div/div[2]/aside/div/input').click()
        driver.find_element_by_xpath('//*[@id="wrapper"]/section[1]/div/div/div[2]/aside/div/input').send_keys(self.metropolitan)
        time.sleep(3)
        driver.find_element_by_xpath('//*[@id="wrapper"]/section[1]/div/div/div[2]/aside/div/button').send_keys(Keys.ENTER)
        print('Connected to Lennar')
        time.sleep(3)
        # Create filter
        print('creating filter')
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[1]/div[1]')))
        time.sleep(3)
        driver.find_element_by_xpath('//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[1]/div[1]').click()
        time.sleep(3)
        # community type
        try:
            driver.find_element_by_xpath('//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[1]/div[3]/div/div[2]/div[2]/div[2]/a').click()
            time.sleep(3)
        except:
            print('no community type')
        # add single family loop
        for i in range(0, 10):
            print(i)
            element = '//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[1]/div[3]/div/div[2]/div[2]/div[2]/div/div/div[2]/div/div/div/div/ul/li[' + str(i) + ']/label'
            try:
                filter = driver.find_element_by_xpath(element).text
            except:
                print('not such element exists')
                filter = 'NO'

            if filter == 'Single Family':
                mainelem = element
                filter = driver.find_element_by_xpath(mainelem)
                filter.click()
                print(filter.text)
                time.sleep(3)
                print('Applied Filters ,success')
                break
            else:
                print('element not found on {}'.format(i))

        # select price
        print('selecting price')
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[1]/div[3]/div/div[2]/div[2]/div[3]/a')))
        driver.find_element_by_xpath('//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[1]/div[3]/div/div[2]/div[2]/div[3]/a').click()
        time.sleep(3)

        try:
            # set price < 300$
            driver.find_element_by_xpath('//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[1]/div[3]/div/div[2]/div[2]/div[3]/div/div/div/div[1]/div[3]/span').click()
            time.sleep(3)
            for i in range(0, 10):
                print(i)
                element = '//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[1]/div[3]/div/div[2]/div[2]/div[3]/div/div/div/div[1]/div[3]/ul/li[' + str(i) + ']'
                print(element)
                try:
                    filter = driver.find_element_by_xpath(element).text
                except:
                    print('not such element exists')
                    filter = 'NO'

                if filter == '300K':  #  300K
                    mainelem = element
                    filter = driver.find_element_by_xpath(mainelem)
                    filter.click()
                    print(filter.text)
                    time.sleep(3)
                    print('Applied Filters ,success')
                    time.sleep(3)
                    print('clicking on botton')
                    driver.find_element_by_xpath('//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[1]/div[3]/div/div[3]/div/div/a[2]').click()
                    break
                else:
                    print('element not found on {}'.format(i))
        except:
            print('price element not found')

        print('locating basic info about communities')
        time.sleep(3)
        try:
            driver = self.driver
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/div[1]/section/div[2]/div/div/div/ul/li[1]/a')))
            self.dict_lennar_filter_info['Communities num'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[1]/section/div[2]/div/div/div/ul/li[1]/a').text
            index1 = self.dict_lennar_filter_info['Communities num'].find('(')
            index2 = self.dict_lennar_filter_info['Communities num'].find(')')
            self.dict_lennar_filter_info['Communities num'] = self.dict_lennar_filter_info['Communities num'][index1 + 1:index2]
            self.dict_lennar_filter_info['Quick Move-In Homes num'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[1]/section/div[2]/div/div/div/ul/li[2]/a').text
            self.dict_lennar_filter_info['Floorplans num'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[1]/section/div[2]/div/div/div/ul/li[3]/a').text
            self.num_of_communities = self.dict_lennar_filter_info['Communities num']
            print('Communities number is {}'.format(self.dict_lennar_filter_info['Communities num']))
            print('change view to list')
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[2]/div/div[3]/a')))
            driver.find_element_by_xpath('//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[2]/div/div[3]/a').click()
            time.sleep(3)
            print('Basic Information scanned {}'.format(self.dict_lennar_filter_info))
        except:
            print('Failed, to Locate information from Lennar')

        try:
            print('copy to xls basic info')
            print('Creating new xls sheet')
            wb = openpyxl.load_workbook(self.xls_name)
            if wb.sheetnames.count(self.metropolitan + ' comm_data') == 0:
                example_sheet = wb['comm_data']
                wb.copy_worksheet(example_sheet)
                new_sheet = wb['comm_data Copy']
                new_sheet.title = self.metropolitan + ' comm_data'
                wb.save(self.xls_name)
                print("XLS new sheet is ready, sheet name: {}".format(new_sheet.title))
                wb.close()
            else:
                print('address was exist in xls')
        except:
            print('failed to connect to xls')

        try:
            time.sleep(3)
            print('opening xls')
            print('xls name {}'.format(self.xls_name))
            wb = openpyxl.load_workbook(self.xls_name)
            sheet = wb[self.metropolitan + ' comm_data']
            sheet['K2'].value = self.dict_lennar_filter_info['metropolitan name']
            sheet['L2'].value = self.dict_lennar_filter_info['Communities num']
            sheet['M2'].value = self.dict_lennar_filter_info['Quick Move-In Homes num']
            sheet['N2'].value = self.dict_lennar_filter_info['Floorplans num']
            sheet['J2'].value = datetime.datetime.now()
            wb.save(self.xls_name)
            wb.close()
            print('sheet name is {}'.format(self.metropolitan + ' comm_data'))
            print('basic community info bar was saved in xls')
            return True
        except:
            print('failed to copy basic community info to XLS ')
            logging.debug('failed to open XLS')
            return False

    def community_and_homes_all_data_to_xls_and_SQL(self):
        try:
            print('Calculating the num of Pages to scroll - communities')
            if int(self.num_of_communities) < 30:
                self.num_of_comm_pages = 1
                print('Num of communities {}'.format(self.num_of_communities))
                print('Num of pages of communities {}'.format(self.num_of_comm_pages))
            else:
                if int(self.num_of_communities) < 60:
                    self.num_of_comm_pages = 2
                    print('Num of communities {}'.format(self.num_of_communities))
                    print('Num of pages of communities {}'.format(self.num_of_comm_pages))
                else:
                    self.num_of_comm_pages = int(self.num_of_communities) / 30
                    self.num_of_comm_pages = round(self.num_of_comm_pages)
                    print('Num of communities {}'.format(self.num_of_communities))
                    print('Num of pages of communities {}'.format(self.num_of_comm_pages))
        except:
            print('could not calculate data about communities')

        if int(self.num_of_communities) < 30:  # if communities < 30 (one page)
            print('if communities < 30')
            time.sleep(2)
            for x in range(0, int(self.num_of_communities)):  # int(self.num_of_communities):
                print('community area entered')
                try:
                    driver = self.driver
                    time.sleep(5)
                    print('change view to list')
                    driver.find_element_by_xpath('//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[2]/div/div[3]/a').click()
                    time.sleep(5)
                except:
                    print('list button not located')
                try:
                    driver = self.driver
                    print('Preparing to Enter community on num {}'.format(x))
                    time.sleep(10)
                    print('trying to locate community address')
                    x_path = '//*[@id="wrapper"]/div[1]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(x + 1) + ']/div[3]/p[2]'
                    print(x_path)
                    self.addr = driver.find_element_by_xpath(x_path).text
                    self.dict_community_data['address'] = self.addr
                    print('Community Address: {}'.format(self.addr))
                    self.x_path_name = '//*[@id="wrapper"]/div[1]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(x + 1) + ']/div[3]/p[1]/a/strong'
                    self.name = driver.find_element_by_xpath(self.x_path_name).text
                    self.dict_community_data['name_community'] = self.name
                    print('Community Name: {}'.format(self.name))
                    self.community_address_list_full.append(self.dict_community_data['address'])
                    print('Community address was added to list for automation')
                    print('scrolling')
                    scroll = 245 * x
                    print(scroll)
                    scroll = "window.scrollTo(0, " + str(scroll) + ")"
                    driver.execute_script(scroll)
                    time.sleep(10)
                    print('scrolled'.format(x))
                    print('trying to click the scrolled community')
                    print(driver.current_url)
                    print(self.x_path_name)
                    print('clicking')
                    driver.find_element_by_xpath(self.x_path_name).click()
                    time.sleep(3)
                    print('clicked')
                    time.sleep(10)
                    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                    print('SUCCESS - community found and pressed')
                    time.sleep(10)
                    c = 0
                except:
                    print('FAILED - to locate community on xpath num {}'.format(x))
                    c = 1

                # if community is located
                if c == 0:
                    print('After Community was located - starting to download data')
                    print('1 First generating ID for Community')
                    time.sleep(5)
                    self.dict_community_data['id_generated'] = uuid.uuid1().int >> 64
                    self.id_random_list.append(self.dict_community_data['id_generated'])
                    print("the id Generated for community is {}".format(self.dict_community_data['id_generated']))

                    try:
                        print('2 try copy overview data')
                        driver = self.driver
                        self.dict_community_data['overview'] = driver.find_element_by_xpath('//*[@id="wrapper"]/section[5]/div/div[2]/div/div[1]/div/div[1]/div[1]/div[2]/div').text
                    except:
                        print('failed to locate overview')

                    time.sleep(5)
                    try:
                        print('3 try copy picture 1 map ')
                        driver = self.driver
                        self.dict_community_data['community_map_url'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[4]/div[2]/div[2]/div[1]/img').get_attribute('src')
                        print('actually downloading the image and changing the name.jpg')
                        urllib.request.urlretrieve(self.dict_community_data['community_map_url'], str(self.dict_community_data['address']) + "_map.jpg")
                    except:
                        print('failed to locate pictures map')
                        self.dict_community_data['community_map_url'] = 'NA'
                    try:
                        print('4 try copy pictures 2')
                        driver = self.driver
                        self.dict_community_data['community_home_picture_for_present_url'] = driver.find_element_by_xpath('//*[@id="tns1"]/div[6]/picture/img').get_attribute('src')
                        urllib.request.urlretrieve(self.dict_community_data['community_home_picture_for_present_url'], str(self.dict_community_data['address']) + "_home_pic.jpg")
                    except:
                        print('failed to locate pictures 2')
                        self.dict_community_data['community_home_picture_for_present_url'] = 'NA'

                    try:
                        print('5 Available Homes and floorplans')
                        driver = self.driver
                        self.dict_community_data['available_homes_quick_move_in_homes'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[2]/div/div/div/ul/li[2]/a').text
                        self.dict_community_data['available_homes_floorplans'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[2]/div/div/div/ul/li[1]/a').text
                        print('success to copy home toolbar data {}'.format(self.dict_community_data))
                    except:
                        print('failed to locate homes toolbar')

                    try:
                        print('6 copy community data to xls')
                        print('open xls '.format(self.xls_name))
                        wb = openpyxl.load_workbook(self.xls_name)
                        time.sleep(2)
                        sheet = wb[self.metropolitan + ' comm_data']
                        sheet['A' + str(self.row)].value = self.dict_community_data['id_generated']
                        sheet['B' + str(self.row)].value = self.dict_community_data['address']
                        sheet['C' + str(self.row)].value = self.dict_community_data['name_community']
                        sheet['D' + str(self.row)].value = self.dict_community_data['overview']
                        sheet['E' + str(self.row)].value = self.dict_community_data['included_features_pdf_url']
                        sheet['F' + str(self.row)].value = self.dict_community_data['community_map_url']
                        sheet['G' + str(self.row)].value = self.dict_community_data['community_home_picture_for_present_url']
                        sheet['H' + str(self.row)].value = self.dict_community_data['available_homes_quick_move_in_homes']
                        sheet['I' + str(self.row)].value = self.dict_community_data['available_homes_floorplans']
                        wb.save(self.xls_name)
                        wb.close()
                        print('COMMUNITY DATA - saved in xls')
                        self.row = self.row + 1
                    except:
                        print('failed to copy community data to XLS ')

                    # homes general data
                    try:
                        driver = self.driver
                        print('scrolling to homes')
                        driver.execute_script("window.scrollTo(0, 2050)")
                        time.sleep(5)
                        print('changing view to list')
                        driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/div/div[1]/div/div/div[2]/div/div[3]/a').click()
                        time.sleep(5)
                        print('Calculating num of homes')
                        self.num_of_moving_homes = self.dict_community_data['available_homes_quick_move_in_homes'][-2:-1]
                        print('num of homes to verify {}'.format(self.num_of_moving_homes))
                        print('num of floorplans to verify {}'.format(self.dict_community_data['available_homes_floorplans'][12:-1]))
                        time.sleep(3)
                    except:
                        print('could not locate general homes information')

                    print('copy homes + floorpans :):):):):):)')
                    print('FLOORPLANS')
                    for j in range(0, int(self.floorplan_homes)):
                        try:
                            driver = self.driver
                            print('Choosing floorplans Homes')
                            driver.execute_script("window.scrollTo(0, 2050)")
                            time.sleep(3)
                            driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[2]/div/div/div/ul/li[1]/a').click()
                            print('floorplans clicked')
                            time.sleep(5)
                            print('For floorplan - Home number {}'.format(j + 1))
                            time.sleep(3)
                            print('Scrolling to Home')
                            scroll = 2000 + (245 * j)
                            scroll = "window.scrollTo(0, " + str(scroll) + ")"
                            driver.execute_script(scroll)
                            print('scrolled to floorplans Home')
                            time.sleep(3)
                            self.dict_home_data['gallery_view_picture'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[1]/div[1]/a[1]/img').get_attribute('src')
                            urllib.request.urlretrieve(self.dict_home_data['gallery_view_picture'], str(self.dict_home_data['home_name']) + ".jpg")
                        except:
                            print('could not locate floorplan home!')

                        print('floorplans - trying to enter - Homes')

                        if int(self.floorplan_homes) <= 1:
                            try:
                                driver = self.driver
                                print('if floorplans home is <= 1 , trying to find home link')
                                print('clicking on floorplans home link')
                                driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div/div[3]/p[1]/a[1]/strong').click()
                                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                                print('floorplans home link clicked')
                                time.sleep(5)
                                print('floorplans home entered')
                                print('waiting for the floorplans home info to appear')
                                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                                print('floorplans Home LOCATED in the list!')
                                time.sleep(15)
                                print('first generating ID floorplans for home')
                                time.sleep(5)
                                print('generating home floorplans id')
                                self.dict_home_data['id_generated_home'] = uuid.uuid1().int >> 64
                                print(type(self.dict_home_data['id_generated_home']))
                                print("the id Generated for floorplans home is {}".format(self.dict_home_data['id_generated_home']))
                            except:
                                print('floorplans could not locate home link <= 1')
                        else:
                            try:
                                driver = self.driver
                                print('IF floorplans Homes count more than > 1')
                                print('clicking on floorplans home link')
                                ActionChains(driver).move_to_element(driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(j + 1) + ']/div[3]/p[1]/a[1]/strong')).perform()
                                time.sleep(5)
                                print('floorplans home name is {}'.format(driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(j + 1) + ']/div[3]/p[1]/a[1]/strong').text))
                                driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(j + 1) + ']/div[3]/p[1]/a[1]/strong').click()
                                print('floorplans home link clicked')
                                time.sleep(5)
                                print('floorplans waiting for the home info to appear')
                                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                                print('Home floorplan LOCATED!')
                                time.sleep(15)
                                print('generating floorplans home id')
                                self.dict_home_data['id_generated_home'] = uuid.uuid1().int >> 64
                                print(type(self.dict_home_data['id_generated_home']))
                                print("the id Generated for floorplans home is {}".format(self.dict_home_data['id_generated_home']))
                            except:
                                print('floorplans Home not located on path number {}'.format(j + 1))

                        try:
                            self.dict_home_data['id_generated'] = self.dict_community_data['id_generated']
                            self.dict_home_data['type'] = "TBB"
                            print('generated id taken from community {}'.format(self.dict_home_data['id_generated']))
                            time.sleep(2)
                        except:
                            print('failed to generate')

                        try:
                            try:
                                driver = self.driver
                                self.dict_home_data['home_name'] = driver.find_element_by_xpath('//*[@id="wrapper"]/section[4]/div/h1').text
                                print(self.dict_home_data['home_name'])
                            except:
                                print('home name not found')

                            try:
                                self.dict_home_data['address'] = self.dict_community_data['address']
                                print(self.dict_home_data['address'])
                            except:
                                print('address not found')

                            try:
                                self.dict_home_data['name_community'] = self.dict_community_data['name_community']
                                print(self.dict_home_data['name_community'])
                            except:
                                print('name community not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['home_site'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[2]').text
                                print(self.dict_home_data['home_site'])
                            except:
                                print('home site not found')

                            self.dict_home_data['included_features_pdf'] = 'under solution'

                            try:
                                self.dict_home_data['availability'] = 'NA'
                                print(self.dict_home_data['availability'])
                            except:
                                print('availability not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['priced_from'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[1]').text
                                self.dict_home_data['priced_from'] = self.dict_home_data['priced_from'][12:-12]
                                print(self.dict_home_data['priced_from'])
                            except:
                                print('priced from not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['home_size'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[2]').text
                                print(self.dict_home_data['home_size'])
                            except:
                                print('home size not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['stories'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[3]').text
                                print(self.dict_home_data['stories'])
                            except:
                                print('stories not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['beds'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[4]').text
                                print(self.dict_home_data['beds'])
                            except:
                                print('beds not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['baths'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[5]').text
                                print(self.dict_home_data['baths'])  # ///*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[4]
                            except:
                                print('baths not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['garage'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[6]').text
                                print(self.dict_home_data['garage'])
                            except:
                                print('garage not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['description'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[2]/div[1]/div/p').text
                                print(self.dict_home_data['description'])
                            except:
                                print('description not found')

                            try:
                                driver = self.driver
                                print('trying to copy home FloorPlan Pic scrolling')
                                driver.execute_script("window.scrollTo(0, 1600)")
                                time.sleep(4)
                                driver.find_element_by_xpath('//*[@id="wrapper"]/div[4]/div/ul/li[2]').click()
                                time.sleep(3)
                                self.dict_home_data['floorplans_with_furniture_pic'] = driver.find_element_by_xpath('//*[@id="tns2-item0"]/div/a/img').get_attribute('src')
                            except:
                                print('could not locate home pics and FloorPlan Pic')

                            # print('Home num {} & Data is: {}'.format(j, self.dict_home_data))
                        except:
                            print('could not locate HOME / elements')

                        print('Trying to copy all gained Homes data to XLS file')
                        try:
                            print('xls - creating new sheet with home name')
                            wb = openpyxl.load_workbook(self.xls_name)
                            if wb.sheetnames.count(self.metropolitan + ' home_data') == 0:
                                print('creating xls')
                                example_sheet = wb['home_data']
                                wb.copy_worksheet(example_sheet)
                                print(wb.sheetnames)
                                new_sheet = wb['home_data Copy']
                                new_sheet.title = self.metropolitan + ' home_data'
                                wb.save(self.xls_name)
                                print("xls new sheet is ready {}".format(self.metropolitan + ' home_data'))
                                print(wb.sheetnames)
                                wb.close()
                            else:
                                print("Metropolitan Homes sheet already created in xls")
                        except:
                            print('failed to connect to xls file and create sheet')

                        # copy home basic info to xls
                        try:
                            # opening xls
                            print('IMPORTANT - copy home info to xls')
                            wb = openpyxl.load_workbook(self.xls_name)
                            sheet = wb[self.metropolitan + ' home_data']
                            sheet['A' + str(self.rowhome)].value = self.dict_home_data['id_generated']
                            sheet['B' + str(self.rowhome)].value = self.dict_home_data['address']
                            sheet['C' + str(self.rowhome)].value = self.dict_home_data['name_community']
                            sheet['D' + str(self.rowhome)].value = self.dict_home_data['home_name']
                            sheet['E' + str(self.rowhome)].value = self.dict_home_data['home_site']
                            sheet['F' + str(self.rowhome)].value = self.dict_home_data['availability']
                            sheet['G' + str(self.rowhome)].value = self.dict_home_data['priced_from']
                            sheet['H' + str(self.rowhome)].value = self.dict_home_data['home_size']
                            sheet['I' + str(self.rowhome)].value = self.dict_home_data['stories']
                            sheet['J' + str(self.rowhome)].value = self.dict_home_data['beds']
                            sheet['K' + str(self.rowhome)].value = self.dict_home_data['baths']
                            sheet['L' + str(self.rowhome)].value = self.dict_home_data['garage']
                            sheet['M' + str(self.rowhome)].value = self.dict_home_data['description']
                            sheet['N' + str(self.rowhome)].value = self.dict_home_data['included_features_pdf']
                            sheet['O' + str(self.rowhome)].value = self.dict_home_data['floorplans_with_furniture_pic']
                            sheet['P' + str(self.rowhome)].value = self.dict_home_data['gallery_view_picture']
                            sheet['R' + str(self.rowhome)].value = self.dict_home_data['type']
                            sheet['Q' + str(self.rowhome)].value = datetime.datetime.now()
                            sheet['S' + str(self.rowhome)].value = self.dict_home_data['id_generated_home']

                            wb.save(self.xls_name)
                            wb.close()
                            print('xls floorplan - HOME params was saved')
                            self.rowhome = self.rowhome + 1
                        except:
                            print('failed to copy floorplans HOME params to xls')
                            logging.debug('failed to open XLS')

                        print('Trying to Connect and copy same data to MySQL server')
                        self.dict_home_data['id_generated_home'] = str(self.dict_home_data['id_generated_home'])
                        try:
                            db = mysql.connector.connect(
                                host='107.180.21.18',
                                user='grow097365',
                                passwd='Jknm678##Tg',
                                database='equity_property'
                            )
                            mycursor = db.cursor()
                            print(db)  # checking our connection to DB
                            command = "SELECT * FROM Limited_Information WHERE id_generated_home = " + "'" + self.dict_home_data['id_generated_home'] + "'"
                            print(command)

                            mycursor.execute(command)
                            myresult = mycursor.fetchall()  # Note: We use the fetchall() method, which fetches all rows from the last executed statement.
                            print(len(myresult))
                            print(myresult)

                            if len(myresult) == 0:
                                print('Similar homes not found, copying to database!')
                                db = mysql.connector.connect(
                                    host='107.180.21.18',
                                    user='grow097365',
                                    passwd='Jknm678##Tg',
                                    database='equity_property'
                                )
                                mycursor = db.cursor()
                                print(db)
                                sql = "INSERT INTO Limited_Information (id_generated, time, address, state, metro, model, size, bedrooms, bathrooms, garage, price, picture_url, type, id_generated_home, name_community) VALUES (%s,%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                                val = (self.dict_home_data['id_generated'],
                                       datetime.datetime.now(),
                                       self.dict_home_data['address'],
                                       self.short_state,
                                       self.metropolitan,
                                       self.dict_home_data['home_name'],
                                       self.dict_home_data['home_size'],
                                       self.dict_home_data['beds'],
                                       self.dict_home_data['baths'],
                                       self.dict_home_data['garage'],
                                       self.dict_home_data['priced_from'],
                                       self.dict_home_data['gallery_view_picture'],
                                       self.dict_home_data['type'],
                                       str(self.dict_home_data['id_generated_home']),
                                       self.dict_home_data['name_community'])
                                mycursor.execute(sql, val)
                                db.commit()
                                time.sleep(3)
                                print('IMPORTANT - Home floorplan data copied to mySQL')
                            else:
                                print('Similar home found in database')
                        except:
                            print('failed to work with mySQL')

                        try:
                            driver = self.driver
                            print('trying to go back to HOMES list after data copied')
                            driver.back()
                            time.sleep(7)
                        except:
                            print('could not go back on general HOMES list')
                    print('HOMES')
                    for j in range(0, int(self.num_of_moving_homes)):
                        try:
                            driver = self.driver
                            print('entering Homes and copy the data')
                            print('For Home number {}'.format(j + 1))
                            print('Choosing quick mov in Homes')
                            driver.execute_script("window.scrollTo(0, 2050)")
                            time.sleep(3)
                            driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[2]/div/div/div/ul/li[2]/a').click()
                            time.sleep(3)
                            print('Scrolling to Home')
                            time.sleep(3)
                            scroll = 2000 + (245 * j)
                            scroll = "window.scrollTo(0, " + str(scroll) + ")"
                            driver.execute_script(scroll)
                            print('scrolled to Homes')
                            print('trying to enter - Homes')
                        except:
                            print('could not locate floorplan home!')

                        if int(self.num_of_moving_homes) <= 1:
                            try:
                                driver = self.driver
                                print('if home is <= 1 , trying to find home link')
                                print('clicking on home link')  # //*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div/div[3]/p[1]/a[1]/strong
                                driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div/div[3]/p[1]/a[1]/strong').click()
                                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                                print('home link clicked')
                                time.sleep(5)
                                print('home entered')
                                print('waiting for the home info to appear')
                                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                                print('Home LOCATED!')
                                time.sleep(15)
                                print('generating home id')
                                self.dict_home_data['id_generated_home'] = uuid.uuid1().int >> 64
                                print(type(self.dict_home_data['id_generated_home']))
                                print("the id Generated for home is {}".format(self.dict_home_data['id_generated_home']))
                            except:
                                print('could not locate home')
                        else:
                            try:
                                driver = self.driver
                                print('if Homes more then > 1')
                                print('clicking on home link')
                                ActionChains(driver).move_to_element(driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(j + 1) + ']/div[3]/p[1]/a[1]/strong')).perform()
                                time.sleep(5)
                                driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(j + 1) + ']/div[3]/p[1]/a[1]/strong').click()
                                time.sleep(5)
                                print('home link clicked')
                                print('home name is {}'.format(driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(j + 1) + ']/div[3]/p[1]/a[1]/strong').text))
                                print('waiting for the home info to appear')
                                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                                print('Home LOCATED!')
                                time.sleep(15)
                                print('generating home id')
                                self.dict_home_data['id_generated_home'] = uuid.uuid1().int >> 64
                                print(type(self.dict_home_data['id_generated_home']))
                                print("the id Generated for home is {}".format(self.dict_home_data['id_generated_home']))
                            except:
                                print('Home not located on path number {}'.format(j + 1))
                        try:
                            self.dict_home_data['id_generated'] = self.dict_community_data['id_generated']
                            self.dict_home_data['type'] = "MIR"
                            print('generated id taken from community {}'.format(self.dict_home_data['id_generated']))

                            time.sleep(2)
                            try:
                                print('try copy home picture')
                                driver = self.driver
                                print('getting the source link of the picture')
                                self.dict_home_data['gallery_view_picture'] = driver.find_element_by_xpath('//*[@id="tns3-item0"]/picture/img').get_attribute('src')
                                print(self.dict_home_data['gallery_view_picture'])
                                urllib.request.urlretrieve(self.dict_home_data['gallery_view_picture'], str(self.dict_home_data['home_name']) + ".jpg")
                            except:
                                print('failed to locate pictures')
                                self.dict_home_data['gallery_view_picture'] = 'NA'

                            try:
                                driver = self.driver
                                self.dict_home_data['home_name'] = driver.find_element_by_xpath('//*[@id="wrapper"]/section[4]/div/h1').text
                                print(self.dict_home_data['home_name'])
                            except:
                                print('home name not found')

                            try:
                                self.dict_home_data['address'] = self.dict_community_data['address']
                                print(self.dict_home_data['address'])
                            except:
                                print('address not found')

                            try:
                                self.dict_home_data['name_community'] = self.dict_community_data['name_community']
                                print(self.dict_home_data['name_community'])
                            except:
                                print('name community not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['home_site'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[1]').text
                                print(self.dict_home_data['home_site'])
                            except:
                                print('home site not found')

                            self.dict_home_data['included_features_pdf'] = 'under solution'

                            try:
                                driver = self.driver
                                self.dict_home_data['availability'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[2]').text
                                print(self.dict_home_data['availability'])
                            except:
                                print('availability not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['priced_from'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[3]').text
                                self.dict_home_data['priced_from'] = self.dict_home_data['priced_from'][12:-12]
                                print(self.dict_home_data['priced_from'])
                            except:
                                print('priced from not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['home_size'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[4]').text
                                print(self.dict_home_data['home_size'])
                            except:
                                print('home size not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['stories'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[5]').text
                                print(self.dict_home_data['stories'])
                            except:
                                print('stories not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['beds'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[6]').text
                                print(self.dict_home_data['beds'])
                            except:
                                print('beds not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['baths'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[7]').text
                                print(self.dict_home_data['baths'])
                            except:
                                print('baths not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['garage'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[8]').text
                                print(self.dict_home_data['garage'])
                            except:
                                print('garage not found')

                            try:
                                driver = self.driver
                                self.dict_home_data['description'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[2]/div[1]/div/p').text
                                print(self.dict_home_data['description'])
                            except:
                                print('description not found')

                            try:
                                driver = self.driver
                                print('trying to copy home FloorPlan Pic scrolling')
                                driver.execute_script("window.scrollTo(0, 1600)")
                                time.sleep(4)
                                driver.find_element_by_xpath('//*[@id="wrapper"]/div[4]/div/ul/li[2]').click()
                                time.sleep(3)
                                self.dict_home_data['floorplans_with_furniture_pic'] = driver.find_element_by_xpath('//*[@id="tns2-item0"]/div/a/img').get_attribute('src')
                            except:
                                print('could not locate home pics and FloorPlan Pic')
                        except:
                            print('could not locate HOME / elements')
                    print('after all homes was scanned, we going back to community')
                    try:
                        driver = self.driver
                        time.sleep(5)
                        driver.back()
                        time.sleep(5)
                        driver.back()
                        time.sleep(10)
                        print('Waiting till the page will load the community')
                    except:
                        print('could not go back on community list')
            print('END of work on communities < 30')
        else:
            print('if communities > 30 and we got pages to scroll')
            time.sleep(2)
            for page in range(self.num_of_comm_pages):
                print('Comm page num {}'.format(page + 1))
                for self.row_num_xls in range(0, 29):  # 30 communities per page
                    print('community area entered')
                    try:
                        driver = self.driver
                        print('change view to list')
                        time.sleep(5)
                        driver.find_element_by_xpath('//*[@id="wrapper"]/div[1]/div/div[1]/div/div/div[2]/div/div[3]/a').click()
                        time.sleep(5)
                    except:
                        print('list button not located')
                    try:
                        driver = self.driver
                        print('Preparing to Enter community on num {}'.format(self.row_num_xls))
                        time.sleep(10)
                        print('trying to locate community address')
                        x_path = '//*[@id="wrapper"]/div[1]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(self.row_num_xls + 1) + ']/div[3]/p[2]'
                        print(x_path)
                        self.addr = driver.find_element_by_xpath(x_path).text
                        self.dict_community_data['address'] = self.addr
                        print('Community Address: {}'.format(self.addr))
                        self.x_path_name = '//*[@id="wrapper"]/div[1]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(self.row_num_xls + 1) + ']/div[3]/p[1]/a/strong'
                        self.name = driver.find_element_by_xpath(self.x_path_name).text
                        self.dict_community_data['name_community'] = self.name
                        print('Community Name: {}'.format(self.name))
                        self.community_address_list_full.append(self.dict_community_data['address'])
                        print('Community address was added to list for automation')

                        print('scrolling')
                        scroll = 245 * self.row_num_xls
                        print(scroll)
                        scroll = "window.scrollTo(0, " + str(scroll) + ")"
                        driver.execute_script(scroll)
                        time.sleep(10)
                        print('scrolled'.format(self.row_num_xls))

                        print('trying to click the scrolled community')
                        print(driver.current_url)
                        print(self.x_path_name)
                        print('clicking')
                        driver.find_element_by_xpath(self.x_path_name).click()
                        time.sleep(3)
                        print('clicked')
                        time.sleep(10)
                        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                        print('SUCCESS - community found and pressed')
                        time.sleep(10)
                        c = 0
                    except:
                        print('FAILED - to locate community on xpath num {}'.format(self.row_num_xls))
                        c = 1

                    # if community is located
                    if c == 0:
                        print('After Community was located - starting to download data')
                        print('1 First generating ID for Community')
                        time.sleep(5)
                        self.dict_community_data['id_generated'] = uuid.uuid1().int >> 64
                        self.id_random_list.append(self.dict_community_data['id_generated'])
                        print("the id Generated for community is {}".format(self.dict_community_data['id_generated']))

                        try:
                            print('2 try copy overview data')
                            driver = self.driver
                            self.dict_community_data['overview'] = driver.find_element_by_xpath('//*[@id="wrapper"]/section[5]/div/div[2]/div/div[1]/div/div[1]/div[1]/div[2]/div').text
                        except:
                            print('failed to locate overview')

                        time.sleep(5)
                        try:
                            print('3 try copy picture 1 map ')
                            driver = self.driver
                            self.dict_community_data['community_map_url'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[4]/div[2]/div[2]/div[1]/img').get_attribute('src')
                            print('actually downloading the image and changing the name.jpg')
                            urllib.request.urlretrieve(self.dict_community_data['community_map_url'], str(self.dict_community_data['address']) + "_map.jpg")
                        except:
                            print('failed to locate pictures map')
                            self.dict_community_data['community_map_url'] = 'NA'
                        try:
                            print('4 try copy pictures 2')
                            driver = self.driver
                            self.dict_community_data['community_home_picture_for_present_url'] = driver.find_element_by_xpath('//*[@id="tns1"]/div[6]/picture/img').get_attribute('src')
                            urllib.request.urlretrieve(self.dict_community_data['community_home_picture_for_present_url'], str(self.dict_community_data['address']) + "_home_pic.jpg")
                        except:
                            print('failed to locate pictures 2')
                            self.dict_community_data['community_home_picture_for_present_url'] = 'NA'

                        try:
                            print('5 Available Homes and floorplans')
                            driver = self.driver
                            self.dict_community_data['available_homes_quick_move_in_homes'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[2]/div/div/div/ul/li[2]/a').text
                            self.dict_community_data['available_homes_floorplans'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[2]/div/div/div/ul/li[1]/a').text
                                                                                                                 # //*[@id="wrapper"]/div[3]/section/div[2]/div/div/div/ul/li[1]
                            print('success to copy home toolbar data {}'.format(self.dict_community_data))
                        except:
                            print('failed to locate homes toolbar')

                        try:
                            print('6 copy community data to xls')
                            print('open xls '.format(self.xls_name))
                            wb = openpyxl.load_workbook(self.xls_name)
                            time.sleep(2)
                            sheet = wb[self.metropolitan + ' comm_data']
                            sheet['A' + str(self.row)].value = self.dict_community_data['id_generated']
                            sheet['B' + str(self.row)].value = self.dict_community_data['address']
                            sheet['C' + str(self.row)].value = self.dict_community_data['name_community']
                            sheet['D' + str(self.row)].value = self.dict_community_data['overview']
                            sheet['E' + str(self.row)].value = self.dict_community_data['included_features_pdf_url']
                            sheet['F' + str(self.row)].value = self.dict_community_data['community_map_url']
                            sheet['G' + str(self.row)].value = self.dict_community_data['community_home_picture_for_present_url']
                            sheet['H' + str(self.row)].value = self.dict_community_data['available_homes_quick_move_in_homes']
                            sheet['I' + str(self.row)].value = self.dict_community_data['available_homes_floorplans']
                            wb.save(self.xls_name)
                            wb.close()
                            print('COMMUNITY DATA - saved in xls')
                            self.row = self.row + 1
                        except:
                            print('failed to copy community data to XLS ')

                        try:
                            print('# homes general data')
                            driver = self.driver
                            print('scrolling to homes')
                            driver.execute_script("window.scrollTo(0, 2050)")
                            time.sleep(5)
                            print('changing view to list')
                            driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/div/div[1]/div/div/div[2]/div/div[3]/a').click()
                            time.sleep(5)
                            print('Calculating num of homes')
                            self.num_of_moving_homes = self.dict_community_data['available_homes_quick_move_in_homes'][-2:-1]
                            self.floorplan_homes = self.dict_community_data['available_homes_floorplans'][-2:-1]
                            print('num of homes to verify {}'.format(self.num_of_moving_homes))
                            print('num of floorplans to verify {}'.format(self.dict_community_data['available_homes_floorplans']))
                            time.sleep(3)
                        except:
                            print('could not locate general homes information')

                        print('copy homes + floorpans :):):):):):)')
                        # print('FLOORPLANS')
                        # for j in range(0, int(self.floorplan_homes)):
                        #     try:
                        #         driver = self.driver
                        #         print('Choosing floorplans Homes')
                        #         driver.execute_script("window.scrollTo(0, 2050)")
                        #         time.sleep(3)
                        #         driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[2]/div/div/div/ul/li[1]/a').click()
                        #         print('floorplans clicked')
                        #         time.sleep(5)
                        #         print('For floorplan - Home number {}'.format(j + 1))
                        #         time.sleep(3)
                        #         print('Scrolling to Home')
                        #         scroll = 2000 + (245 * j)
                        #         scroll = "window.scrollTo(0, " + str(scroll) + ")"
                        #         driver.execute_script(scroll)
                        #         print('scrolled to floorplans Home')
                        #         time.sleep(3)
                        #         self.dict_home_data['gallery_view_picture'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[1]/div[1]/a[1]/img').get_attribute('src')
                        #         urllib.request.urlretrieve(self.dict_home_data['gallery_view_picture'], str(self.dict_home_data['home_name']) + ".jpg")
                        #     except:
                        #         print('could not locate floorplan home!')
                        #
                        #     print('floorplans - trying to enter - Homes')
                        #
                        #     if int(self.floorplan_homes) <= 1:
                        #         try:
                        #             driver = self.driver
                        #             print('if floorplans home is <= 1 , trying to find home link')
                        #             print('clicking on floorplans home link')
                        #             driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div/div[3]/p[1]/a[1]/strong').click()
                        #             WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                        #             print('floorplans home link clicked')
                        #             time.sleep(5)
                        #             print('floorplans home entered')
                        #             print('waiting for the floorplans home info to appear')
                        #             WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                        #             print('floorplans Home LOCATED in the list!')
                        #             time.sleep(15)
                        #             print('first generating ID floorplans for home')
                        #             time.sleep(5)
                        #             print('generating home floorplans id')
                        #             self.dict_home_data['id_generated_home'] = uuid.uuid1().int >> 64
                        #             print(type(self.dict_home_data['id_generated_home']))
                        #             print("the id Generated for floorplans home is {}".format(self.dict_home_data['id_generated_home']))
                        #         except:
                        #             print('floorplans could not locate home link <= 1')
                        #     else:
                        #         try:
                        #             driver = self.driver
                        #             print('IF floorplans Homes count more than > 1')
                        #             print('clicking on floorplans home link')
                        #             ActionChains(driver).move_to_element(driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(j + 1) + ']/div[3]/p[1]/a[1]/strong')).perform()
                        #             time.sleep(5)
                        #             print('floorplans home name is {}'.format(driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(j + 1) + ']/div[3]/p[1]/a[1]/strong').text))
                        #             driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(j + 1) + ']/div[3]/p[1]/a[1]/strong').click()
                        #             print('floorplans home link clicked')
                        #             time.sleep(5)
                        #             print('floorplans waiting for the home info to appear')
                        #             WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                        #             print('Home floorplan LOCATED!')
                        #             time.sleep(15)
                        #             print('generating floorplans home id')
                        #             self.dict_home_data['id_generated_home'] = uuid.uuid1().int >> 64
                        #             print(type(self.dict_home_data['id_generated_home']))
                        #             print("the id Generated for floorplans home is {}".format(self.dict_home_data['id_generated_home']))
                        #         except:
                        #             print('floorplans Home not located on path number {}'.format(j + 1))
                        #
                        #     try:
                        #         self.dict_home_data['id_generated'] = self.dict_community_data['id_generated']
                        #         self.dict_home_data['type'] = "TBB"
                        #         print('generated id taken from community {}'.format(self.dict_home_data['id_generated']))
                        #         time.sleep(2)
                        #     except:
                        #         print('failed to generate')
                        #
                        #     try:
                        #         try:
                        #             driver = self.driver
                        #             self.dict_home_data['home_name'] = driver.find_element_by_xpath('//*[@id="wrapper"]/section[4]/div/h1').text
                        #             print(self.dict_home_data['home_name'])
                        #         except:
                        #             print('home name not found')
                        #
                        #         try:
                        #             self.dict_home_data['address'] = self.dict_community_data['address']
                        #             print(self.dict_home_data['address'])
                        #         except:
                        #             print('address not found')
                        #
                        #         try:
                        #             self.dict_home_data['name_community'] = self.dict_community_data['name_community']
                        #             print(self.dict_home_data['name_community'])
                        #         except:
                        #             print('name community not found')
                        #
                        #         try:
                        #             driver = self.driver
                        #             self.dict_home_data['home_site'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[2]').text
                        #             print(self.dict_home_data['home_site'])
                        #         except:
                        #             print('home site not found')
                        #
                        #         self.dict_home_data['included_features_pdf'] = 'under solution'
                        #
                        #         try:
                        #             self.dict_home_data['availability'] = 'NA'
                        #             print(self.dict_home_data['availability'])
                        #         except:
                        #             print('availability not found')
                        #
                        #         try:
                        #             driver = self.driver
                        #             self.dict_home_data['priced_from'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[1]').text
                        #             self.dict_home_data['priced_from'] = self.dict_home_data['priced_from'][12:-12]
                        #             print(self.dict_home_data['priced_from'])
                        #         except:
                        #             print('priced from not found')
                        #
                        #         try:
                        #             driver = self.driver
                        #             self.dict_home_data['home_size'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[2]').text
                        #             print(self.dict_home_data['home_size'])
                        #         except:
                        #             print('home size not found')
                        #
                        #         try:
                        #             driver = self.driver
                        #             self.dict_home_data['stories'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[3]').text
                        #             print(self.dict_home_data['stories'])
                        #         except:
                        #             print('stories not found')
                        #
                        #         try:
                        #             driver = self.driver
                        #             self.dict_home_data['beds'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[4]').text
                        #             print(self.dict_home_data['beds'])
                        #         except:
                        #             print('beds not found')
                        #
                        #         try:
                        #             driver = self.driver
                        #             self.dict_home_data['baths'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[5]').text
                        #             print(self.dict_home_data['baths'])  # ///*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[4]
                        #         except:
                        #             print('baths not found')
                        #
                        #         try:
                        #             driver = self.driver
                        #             self.dict_home_data['garage'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[6]').text
                        #             print(self.dict_home_data['garage'])
                        #         except:
                        #             print('garage not found')
                        #
                        #         try:
                        #             driver = self.driver
                        #             self.dict_home_data['description'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[2]/div[1]/div/p').text
                        #             print(self.dict_home_data['description'])
                        #         except:
                        #             print('description not found')
                        #
                        #         try:
                        #             driver = self.driver
                        #             print('trying to copy home FloorPlan Pic scrolling')
                        #             driver.execute_script("window.scrollTo(0, 1600)")
                        #             time.sleep(4)
                        #             driver.find_element_by_xpath('//*[@id="wrapper"]/div[4]/div/ul/li[2]').click()
                        #             time.sleep(3)
                        #             self.dict_home_data['floorplans_with_furniture_pic'] = driver.find_element_by_xpath('//*[@id="tns2-item0"]/div/a/img').get_attribute('src')
                        #         except:
                        #             print('could not locate home pics and FloorPlan Pic')
                        #
                        #         # print('Home num {} & Data is: {}'.format(j, self.dict_home_data))
                        #     except:
                        #         print('could not locate HOME / elements')
                        #
                        #     print('Trying to copy all gained Homes data to XLS file')
                        #     try:
                        #         print('xls - creating new sheet with home name')
                        #         wb = openpyxl.load_workbook(self.xls_name)
                        #         if wb.sheetnames.count(self.metropolitan + ' home_data') == 0:
                        #             print('creating xls')
                        #             example_sheet = wb['home_data']
                        #             wb.copy_worksheet(example_sheet)
                        #             print(wb.sheetnames)
                        #             new_sheet = wb['home_data Copy']
                        #             new_sheet.title = self.metropolitan + ' home_data'
                        #             wb.save(self.xls_name)
                        #             print("xls new sheet is ready {}".format(self.metropolitan + ' home_data'))
                        #             print(wb.sheetnames)
                        #             wb.close()
                        #         else:
                        #             print("Metropolitan Homes sheet already created in xls")
                        #     except:
                        #         print('failed to connect to xls file and create sheet')
                        #
                        #     # copy home basic info to xls
                        #     try:
                        #         # opening xls
                        #         print('IMPORTANT - copy home info to xls')
                        #         wb = openpyxl.load_workbook(self.xls_name)
                        #         sheet = wb[self.metropolitan + ' home_data']
                        #         sheet['A' + str(self.rowhome)].value = self.dict_home_data['id_generated']
                        #         sheet['B' + str(self.rowhome)].value = self.dict_home_data['address']
                        #         sheet['C' + str(self.rowhome)].value = self.dict_home_data['name_community']
                        #         sheet['D' + str(self.rowhome)].value = self.dict_home_data['home_name']
                        #         sheet['E' + str(self.rowhome)].value = self.dict_home_data['home_site']
                        #         sheet['F' + str(self.rowhome)].value = self.dict_home_data['availability']
                        #         sheet['G' + str(self.rowhome)].value = self.dict_home_data['priced_from']
                        #         sheet['H' + str(self.rowhome)].value = self.dict_home_data['home_size']
                        #         sheet['I' + str(self.rowhome)].value = self.dict_home_data['stories']
                        #         sheet['J' + str(self.rowhome)].value = self.dict_home_data['beds']
                        #         sheet['K' + str(self.rowhome)].value = self.dict_home_data['baths']
                        #         sheet['L' + str(self.rowhome)].value = self.dict_home_data['garage']
                        #         sheet['M' + str(self.rowhome)].value = self.dict_home_data['description']
                        #         sheet['N' + str(self.rowhome)].value = self.dict_home_data['included_features_pdf']
                        #         sheet['O' + str(self.rowhome)].value = self.dict_home_data['floorplans_with_furniture_pic']
                        #         sheet['P' + str(self.rowhome)].value = self.dict_home_data['gallery_view_picture']
                        #         sheet['R' + str(self.rowhome)].value = self.dict_home_data['type']
                        #         sheet['Q' + str(self.rowhome)].value = datetime.datetime.now()
                        #         sheet['S' + str(self.rowhome)].value = self.dict_home_data['id_generated_home']
                        #
                        #         wb.save(self.xls_name)
                        #         wb.close()
                        #         print('xls floorplan - HOME params was saved')
                        #         self.rowhome = self.rowhome + 1
                        #     except:
                        #         print('failed to copy floorplans HOME params to xls')
                        #         logging.debug('failed to open XLS')
                        #
                        #     print('Trying to Connect and copy same data to MySQL server')
                        #     self.dict_home_data['id_generated_home'] = str(self.dict_home_data['id_generated_home'])
                        #     try:
                        #         db = mysql.connector.connect(
                        #             host='107.180.21.18',
                        #             user='grow097365',
                        #             passwd='Jknm678##Tg',
                        #             database='equity_property'
                        #         )
                        #         mycursor = db.cursor()
                        #         print(db)  # checking our connection to DB
                        #         command = "SELECT * FROM Limited_Information WHERE id_generated_home = " + "'" + self.dict_home_data['id_generated_home'] + "'"
                        #         print(command)
                        #
                        #         mycursor.execute(command)
                        #         myresult = mycursor.fetchall()  # Note: We use the fetchall() method, which fetches all rows from the last executed statement.
                        #         print(len(myresult))
                        #         print(myresult)
                        #
                        #         if len(myresult) == 0:
                        #             print('Similar homes not found, copying to database!')
                        #             db = mysql.connector.connect(
                        #                 host='107.180.21.18',
                        #                 user='grow097365',
                        #                 passwd='Jknm678##Tg',
                        #                 database='equity_property'
                        #             )
                        #             mycursor = db.cursor()
                        #             print(db)
                        #             sql = "INSERT INTO Limited_Information (id_generated, time, address, state, metro, model, size, bedrooms, bathrooms, garage, price, picture_url, type, id_generated_home, name_community) VALUES (%s,%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                        #             val = (self.dict_home_data['id_generated'],
                        #                    datetime.datetime.now(),
                        #                    self.dict_home_data['address'],
                        #                    self.short_state,
                        #                    self.metropolitan,
                        #                    self.dict_home_data['home_name'],
                        #                    self.dict_home_data['home_size'],
                        #                    self.dict_home_data['beds'],
                        #                    self.dict_home_data['baths'],
                        #                    self.dict_home_data['garage'],
                        #                    self.dict_home_data['priced_from'],
                        #                    self.dict_home_data['gallery_view_picture'],
                        #                    self.dict_home_data['type'],
                        #                    str(self.dict_home_data['id_generated_home']),
                        #                    self.dict_home_data['name_community'])
                        #             mycursor.execute(sql, val)
                        #             db.commit()
                        #             time.sleep(3)
                        #             print('IMPORTANT - Home floorplan data copied to mySQL')
                        #         else:
                        #             print('Similar home found in database')
                        #     except:
                        #         print('failed to work with mySQL')
                        #
                        #     try:
                        #         driver = self.driver
                        #         print('trying to go back to HOMES list after data copied')
                        #         driver.back()
                        #         time.sleep(7)
                        #     except:
                        #         print('could not go back on general HOMES list')
                        # print('HOMES')
                        # for j in range(0, int(self.num_of_moving_homes)):
                        #     try:
                        #         driver = self.driver
                        #         print('entering Homes and copy the data')
                        #         print('For Home number {}'.format(j + 1))
                        #         print('Choosing quick mov in Homes')
                        #         driver.execute_script("window.scrollTo(0, 2050)")
                        #         time.sleep(3)
                        #         driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[2]/div/div/div/ul/li[2]/a').click()
                        #         time.sleep(3)
                        #         print('Scrolling to Home')
                        #         time.sleep(3)
                        #         scroll = 2000 + (245 * j)
                        #         scroll = "window.scrollTo(0, " + str(scroll) + ")"
                        #         driver.execute_script(scroll)
                        #         print('scrolled to Homes')
                        #         print('trying to enter - Homes')
                        #     except:
                        #         print('could not locate floorplan home!')
                        #
                        #     if int(self.num_of_moving_homes) <= 1:
                        #         try:
                        #             driver = self.driver
                        #             print('if home is <= 1 , trying to find home link')
                        #             print('clicking on home link')  # //*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div/div[3]/p[1]/a[1]/strong
                        #             driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div/div[3]/p[1]/a[1]/strong').click()
                        #             WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                        #             print('home link clicked')
                        #             time.sleep(5)
                        #             print('home entered')
                        #             print('waiting for the home info to appear')
                        #             WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                        #             print('Home LOCATED!')
                        #             time.sleep(15)
                        #             print('generating home id')
                        #             self.dict_home_data['id_generated_home'] = uuid.uuid1().int >> 64
                        #             print(type(self.dict_home_data['id_generated_home']))
                        #             print("the id Generated for home is {}".format(self.dict_home_data['id_generated_home']))
                        #         except:
                        #             print('could not locate home')
                        #     else:
                        #         try:
                        #             driver = self.driver
                        #             print('if Homes more then > 1')
                        #             print('clicking on home link')
                        #             ActionChains(driver).move_to_element(driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(j + 1) + ']/div[3]/p[1]/a[1]/strong')).perform()
                        #             time.sleep(5)
                        #             driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(j + 1) + ']/div[3]/p[1]/a[1]/strong').click()
                        #             time.sleep(5)
                        #             print('home link clicked')
                        #             print('home name is {}'.format(driver.find_element_by_xpath('//*[@id="wrapper"]/div[3]/section/div[3]/div[2]/div/div[1]/div[2]/div[2]/div[1]/div[' + str(j + 1) + ']/div[3]/p[1]/a[1]/strong').text))
                        #             print('waiting for the home info to appear')
                        #             WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wrapper"]/section[4]/div/h1')))
                        #             print('Home LOCATED!')
                        #             time.sleep(15)
                        #             print('generating home id')
                        #             self.dict_home_data['id_generated_home'] = uuid.uuid1().int >> 64
                        #             print(type(self.dict_home_data['id_generated_home']))
                        #             print("the id Generated for home is {}".format(self.dict_home_data['id_generated_home']))
                        #         except:
                        #             print('Home not located on path number {}'.format(j + 1))
                        #     try:
                        #         self.dict_home_data['id_generated'] = self.dict_community_data['id_generated']
                        #         self.dict_home_data['type'] = "MIR"
                        #         print('generated id taken from community {}'.format(self.dict_home_data['id_generated']))
                        #
                        #         time.sleep(2)
                        #         try:
                        #             print('try copy home picture')
                        #             driver = self.driver
                        #             print('getting the source link of the picture')
                        #             self.dict_home_data['gallery_view_picture'] = driver.find_element_by_xpath('//*[@id="tns3-item0"]/picture/img').get_attribute('src')
                        #             print(self.dict_home_data['gallery_view_picture'])
                        #             urllib.request.urlretrieve(self.dict_home_data['gallery_view_picture'], str(self.dict_home_data['home_name']) + ".jpg")
                        #         except:
                        #             print('failed to locate pictures')
                        #             self.dict_home_data['gallery_view_picture'] = 'NA'
                        #
                        #         try:
                        #             driver = self.driver
                        #             self.dict_home_data['home_name'] = driver.find_element_by_xpath('//*[@id="wrapper"]/section[4]/div/h1').text
                        #             print(self.dict_home_data['home_name'])
                        #         except:
                        #             print('home name not found')
                        #
                        #         try:
                        #             self.dict_home_data['address'] = self.dict_community_data['address']
                        #             print(self.dict_home_data['address'])
                        #         except:
                        #             print('address not found')
                        #
                        #         try:
                        #             self.dict_home_data['name_community'] = self.dict_community_data['name_community']
                        #             print(self.dict_home_data['name_community'])
                        #         except:
                        #             print('name community not found')
                        #
                        #         try:
                        #             driver = self.driver
                        #             self.dict_home_data['home_site'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[1]').text
                        #             print(self.dict_home_data['home_site'])
                        #         except:
                        #             print('home site not found')
                        #
                        #         self.dict_home_data['included_features_pdf'] = 'under solution'
                        #
                        #         try:
                        #             driver = self.driver
                        #             self.dict_home_data['availability'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[2]').text
                        #             print(self.dict_home_data['availability'])
                        #         except:
                        #             print('availability not found')
                        #
                        #         try:
                        #             driver = self.driver
                        #             self.dict_home_data['priced_from'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[3]').text
                        #             self.dict_home_data['priced_from'] = self.dict_home_data['priced_from'][12:-12]
                        #             print(self.dict_home_data['priced_from'])
                        #         except:
                        #             print('priced from not found')
                        #
                        #         try:
                        #             driver = self.driver
                        #             self.dict_home_data['home_size'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[4]').text
                        #             print(self.dict_home_data['home_size'])
                        #         except:
                        #             print('home size not found')
                        #
                        #         try:
                        #             driver = self.driver
                        #             self.dict_home_data['stories'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[5]').text
                        #             print(self.dict_home_data['stories'])
                        #         except:
                        #             print('stories not found')
                        #
                        #         try:
                        #             driver = self.driver
                        #             self.dict_home_data['beds'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[6]').text
                        #             print(self.dict_home_data['beds'])
                        #         except:
                        #             print('beds not found')
                        #
                        #         try:
                        #             driver = self.driver
                        #             self.dict_home_data['baths'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[7]').text
                        #             print(self.dict_home_data['baths'])
                        #         except:
                        #             print('baths not found')
                        #
                        #         try:
                        #             driver = self.driver
                        #             self.dict_home_data['garage'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[1]/ul/li[8]').text
                        #             print(self.dict_home_data['garage'])
                        #         except:
                        #             print('garage not found')
                        #
                        #         try:
                        #             driver = self.driver
                        #             self.dict_home_data['description'] = driver.find_element_by_xpath('//*[@id="wrapper"]/div[2]/div[1]/div[2]/div[1]/div/p').text
                        #             print(self.dict_home_data['description'])
                        #         except:
                        #             print('description not found')
                        #
                        #         try:
                        #             driver = self.driver
                        #             print('trying to copy home FloorPlan Pic scrolling')
                        #             driver.execute_script("window.scrollTo(0, 1600)")
                        #             time.sleep(4)
                        #             driver.find_element_by_xpath('//*[@id="wrapper"]/div[4]/div/ul/li[2]').click()
                        #             time.sleep(3)
                        #             self.dict_home_data['floorplans_with_furniture_pic'] = driver.find_element_by_xpath('//*[@id="tns2-item0"]/div/a/img').get_attribute('src')
                        #         except:
                        #             print('could not locate home pics and FloorPlan Pic')
                        #     except:
                        #         print('could not locate HOME / elements')

                        print('after all homes was scanned, we going back to community')
                        try:
                            driver = self.driver
                            time.sleep(5)
                            driver.back()
                            time.sleep(5)
                            driver.back()
                            time.sleep(10)
                            print('Waiting till the page will load the community')
                        except:
                            print('could not go back on community list')
                    else:
                        print('End of community list, number of communities was {} '.format(self.num_of_communities))
                # page num
                try:
                    driver = self.driver
                    print('scrolling to next page button')
                    driver.execute_script("window.scrollTo(0, 8000)")
                    time.sleep(5)

                    for i in range(10, 0, -1):
                        try:
                            print('trying to locate next page button {}'.format(i))
                            driver.find_element_by_xpath('//*[@id="wrapper"]/div[1]/section/div[3]/div[2]/div/div[2]/a[' + str(i) + ']').click()
                            time.sleep(5)
                            print('next button located on number {}'.format(i))
                        except:
                            print('trying another path - next page button was not located')

                    print('next page pressed')
                    time.sleep(2)
                    driver.execute_script("window.scrollTo(0, 0)")
                    print('scrolling back to top')
                    time.sleep(6)
                except:
                    print('button not located')
            print('END of work on communities > 30')

    def return_community_address_list(self):
        return self.community_address_list_full

    def return_Generated_Id_list(self):
        return self.id_random_list

    def return_list_of_homes(self):
        return self.list_of_homes






