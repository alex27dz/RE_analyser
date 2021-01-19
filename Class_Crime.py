'''
CRIME Class is built for adding the crime information of the given address
Using address to collect information from all source web sites using Beautiful Soup and store it in dictionaries then in MySQL
'''
import openpyxl
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import time
import pprint
import logging
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.action_chains import ActionChains


class Crime(object):
    def __init__(self, street, state, city, short_state, xls_name):
        # all setup params
        self.street = street
        self.state = state
        self.city = city
        self.short_state = short_state
        self.xls_name = xls_name
        self.full_addr = self.street.lower() + " " + self.city.lower() + " " + self.state.lower() + " " + self.short_state.lower()
        self.driver = webdriver.Chrome("/Users/alexdezho/Downloads/chromedriver")

        #urls
        self.onboardnavigator_url = 'http://www.onboardnavigator.com/webcontent/OBWC_Search.aspx?&AID=102'
        self.city_data_url = 'http://www.city-data.com'
        self.home_facts_url = 'https://www.homefacts.com/'
        self.neighborhoodscout_url = 'https://www.neighborhoodscout.com/' + self.short_state.lower() + '/' + self.city.lower() + '/crime'
        self.bestplaces_url = 'https://www.bestplaces.net/crime/city/' + self.state.lower() + '/' + self.city.lower()
        # add NA
        #dictionaries
        self.dict_crime_total = {
            'Crime Index city': 'NA',
            'US avarage': 'NA',
            'Pic of graph': 'NA',
            'total info': 'NA',
            'Overall Score': 'NA',
            'Overall score big num': 'NA',
            'Score small procents': 'NA',
            'Violent crime & US average': 'NA',
            'Property crime & US average': 'NA',
            'Photos and Maps of the city': 'NA',
        }
        self.dict_basic_info = {
            'street': self.street,
            'city': self.city,
            'short_state': self.short_state,
            'state': self.state,
            'zip_code': 'NA',
            'metropolitan': 'NA',
            'link_google_maps': 'NA'
        }
        self.dict_onboardnavigator = {
            'Total personal': 'NA',
            'Total property': 'NA',
            'Total overall': 'NA',
            'Year': '2019',

        }
        self.dict_city_data = {
            'Crime Index city': 'NA',
            'US avarage': 'NA',
            'Pic of graph': 'NA',
            'total info': 'NA',
            'Year': '2019',

        }
        self.dict_home_facts = {
            'Overall Score': 'NA',
            'Overall score big num': 'NA',
            'Score small procents': 'NA',
            'Year': '2019',

        }
        self.dict_offenders = {
            'offender1': 'NA',
            'offender2': 'NA',
            'offender3': 'NA',

        }
        self.dict_neighborhoodscout = {
            'Diagram': 'NA',
            'List of safe areas': 'NA',
        }
        self.dict_bestplaces = {
            'Violent crime & US average': 'NA',
            'Property crime & US average': 'NA',
            'Photos and Maps of the city': 'NA',
        }

    def closeBrowser(self):
        self.driver.close()
        logging.debug('Browser closed')
        print('Browser closed')
# the functions below written in a working flow
# getting all the information and copy into dicts
    def onboardnavigator_to_dict(self):
        try:
            print('onboardnavigator')
            driver = self.driver
            driver.get(self.onboardnavigator_url)
            time.sleep(10)
            print('Navigator tool opened')
            # select state
            state = driver.find_element_by_xpath('//*[@id="ddlGenLookupStateID"]').click()
            time.sleep(5)
            Select(driver.find_element_by_tag_name('select')).select_by_visible_text(self.state)
            time.sleep(5)
            print('state selected')
            driver.find_element_by_xpath('//*[@id="tbGenSearch"]').send_keys(self.city)
            time.sleep(3)
            driver.find_element_by_xpath('//*[@id="radGenCity"]').click()
            time.sleep(3)
            driver.find_element_by_xpath('//*[@id="cmdGenSave"]').click()
            time.sleep(10)
            print('navigator address located')
            link = driver.current_url
            self.dict_onboardnavigator['Total personal'] = link
            self.dict_onboardnavigator['Total property'] = link
            self.dict_onboardnavigator['Total overall'] = link
            print('onboardnavigator params was copied to dictionary , success {}'.format(self.dict_onboardnavigator))
        except:
            print('failed to locate navigator')

    def city_data_to_dict(self):
        try:
            print('citydata')
            driver = self.driver
            driver.get(self.city_data_url)
            time.sleep(10)
            driver.find_element_by_xpath('//*[@id="intelligent_search"]').click()
            time.sleep(3)
            driver.find_element_by_xpath('//*[@id="intelligent_search"]').send_keys(self.city + ' ' + self.state)
            time.sleep(3)
            driver.find_element_by_xpath('//*[@id="search_bar_box"]/input[2]').click()
            time.sleep(10)
            driver.execute_script("window.scrollTo(0,4100)")
            time.sleep(10)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="sex-offenders"]/p')))
            # select city data elemnets and copy to dictionary
            self.dict_city_data['total info'] = driver.find_element_by_xpath('//*[@id="sex-offenders"]/p').text
            self.dict_city_data['Pic of graph'] = driver.current_url
            self.dict_crime_total['total info'] = self.dict_city_data['total info']
            self.dict_crime_total['Pic of graph'] = self.dict_city_data['Pic of graph']
            print('city_data total info copied {}'.format(self.dict_city_data))
        except:
            print('failed to locate city data elements')
        try:
            driver = self.driver
            self.dict_city_data['Crime Index city'] = driver.find_element_by_xpath('//*[@id="crimeTab"]/tfoot/tr/td[15]').text
            self.dict_city_data['US avarage'] = driver.find_element_by_xpath('//*[@id="crimeTab"]/tfoot/tr/td[1]').text
            self.dict_crime_total['Crime Index city'] = self.dict_city_data['Crime Index city']
            self.dict_crime_total['US avarage'] = self.dict_city_data['US avarage']
            print('crime table params was copied to dictionary , success {}'.format(self.dict_city_data))
            logging.debug('crime table params was copied to dictionary , success {}'.format(self.dict_city_data))
            return True
        except:
            self.dict_city_data['Crime Index city'] = 'Crime table not exists in city_data for this state'
            print('Crime table not exists in city_data for this state')
            logging.debug('fail')
            return False

    def home_facts_to_dict(self):
        try:
            print('homefacts')
            driver = self.driver
            driver.get(self.home_facts_url)
            time.sleep(10)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="fulladdress"]')))
            addr = driver.find_element_by_xpath('//*[@id="fulladdress"]')
            addr.click()
            time.sleep(3)
            addr.send_keys(self.full_addr)
            time.sleep(3)
            driver.find_element_by_xpath('//*[@id="main-search-form"]/div/div/div/div[1]/span/button').click()
            time.sleep(10)
            element = driver.find_element_by_xpath('/html/body/section[2]/div[2]/div[2]/div[1]/div[3]/ul/li[1]/span[4]/a')
            driver.execute_script("window.scrollTo(0,600)")
            time.sleep(3)
            element.click()
            print(driver.current_url)
            print('view crime statistics report')
            time.sleep(10)
            try:
                print('trying to click')
                driver.find_element_by_partial_link_text('view crime statistics report').click()
            except:
                print('trying to click with second option')
                driver.execute_script("window.scrollTo(0,2700)")
                time.sleep(7)
                driver.find_element_by_partial_link_text('view crime statistics report').click()

            time.sleep(10)
            print(driver.current_url)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="crimeScore"]/div[1]/div[4]')))
            self.dict_home_facts['Overall Score'] = driver.find_element_by_xpath('//*[@id="crimeScore"]/div[1]/div[4]').get_attribute('class')
            self.dict_home_facts['Overall score big num'] = driver.find_element_by_xpath('//*[@id="crimeScore"]/div[1]/div[2]').text
            self.dict_home_facts['Score small procents'] = driver.find_element_by_xpath('//*[@id="crimeScore"]/div[1]/div[3]').text
            self.dict_home_facts['Overall Score'] = self.dict_home_facts['Overall Score']
            self.dict_home_facts['Overall score big num'] = self.dict_home_facts['Overall score big num']
            self.dict_home_facts['Score small procents'] = self.dict_home_facts['Score small procents']
            print(self.dict_home_facts['Overall Score'])
            print(self.dict_home_facts['Overall score big num'])
            print(self.dict_home_facts['Score small procents'])
            self.dict_crime_total['Overall Score'] = self.dict_home_facts['Overall Score']
            self.dict_crime_total['Overall score big num'] = self.dict_home_facts['Overall score big num']
            self.dict_crime_total['Score small procents'] = self.dict_home_facts['Score small procents']
            print('dict_home_facts params was copied to dictionary , success {}'.format(self.dict_home_facts))
            print('dict_offenders params was copied to dictionary , success {}'.format(self.dict_offenders))
        except:
            print('failed to locate and copy from home facts')

    def neighborhoodscout_to_dict(self):
        try:
            print('neighborhoodscout')
            print(self.neighborhoodscout_url)
            data = requests.get(self.neighborhoodscout_url)
            time.sleep(5)
            soup = BeautifulSoup(data.content, 'html.parser')
            list = soup.find_all('script', type='application/ld+json')
            list = str(list)
            index_list_start = list.find('itemListOrder')
            new_list = list[index_list_start:]
            index_list_end = new_list.find('</script>')
            orig_list = new_list[:index_list_end]
            index1 = orig_list.find('[')
            orig_list = orig_list[index1:]
            index2 = orig_list.find(']')
            # list of safety places taken from HTML converted to string
            orig_list = orig_list[index1:index2]
            self.dict_neighborhoodscout['List of safe areas'] = orig_list
            self.dict_neighborhoodscout['Diagram'] = self.neighborhoodscout_url
            print('neighborhoodscout params was copied to dictionary , success {}'.format(self.dict_neighborhoodscout))
            logging.debug('neighborhoodscout params was copied to dictionary , success {}'.format(self.dict_neighborhoodscout))
            return True
        except:
            logging.debug('fail to connect or copy from neighborhoodscout')
            print('fail to connect or copy from neighborhoodscout')
            return False

    def bestplaces_to_dict(self):
        try:
            print('bestplaces')
            driver = self.driver
            driver.get(self.bestplaces_url)
            time.sleep(5)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="form1"]/div[7]/div[2]/div[2]/div[2]/div/h5[1]')))
            self.dict_bestplaces['Violent crime & US average'] = driver.find_element_by_xpath('//*[@id="form1"]/div[7]/div[2]/div[2]/div[2]/div/h5[1]').text
            self.dict_bestplaces['Property crime & US average'] = driver.find_element_by_xpath('//*[@id="form1"]/div[7]/div[2]/div[2]/div[2]/div/h5[2]').text

            # Photos and Maps
            driver.find_element_by_xpath('//*[@id="form1"]/div[5]/div/div/p[3]/a[3]/u').click()
            time.sleep(4)
            self.dict_bestplaces['Photos and Maps of the city'] = driver.current_url
            print('bestplaces params was copied to dictionary , success {}'.format(self.dict_bestplaces))
            logging.debug('bestplaces params was copied to dictionary , success {}'.format(self.dict_bestplaces))

            self.dict_crime_total['Violent crime & US average'] = self.dict_bestplaces['Violent crime & US average']
            self.dict_crime_total['Property crime & US average'] = self.dict_bestplaces['Property crime & US average']
            self.dict_crime_total['Photos and Maps of the city'] = self.dict_bestplaces['Photos and Maps of the city']

            return True
        except:
            logging.debug('fail to connect or copy from bestplaces')
            print('fail to connect or copy from bestplaces')
            return False

# print all dictionaries
    def printall(self):
        print('All dictionaries\n')
        pp = pprint.PrettyPrinter(indent=4)
        print(self.dict_basic_info)
        pp.pprint(self.dict_basic_info)
        print(self.dict_onboardnavigator)
        pp.pprint(self.dict_onboardnavigator)
        print(self.dict_city_data)
        pp.pprint(self.dict_city_data)
        print(self.dict_home_facts)
        pp.pprint(self.dict_home_facts)
        print(self.dict_offenders)
        pp.pprint(self.dict_offenders)
        print(self.dict_neighborhoodscout)
        pp.pprint(self.dict_neighborhoodscout)
        print(self.dict_bestplaces)
        pp.pprint(self.dict_bestplaces)
        return True
# returning all dictionaries for future use to add to general list
    def return_dict_basic_info(self):
        return self.dict_basic_info

    def return_dict_onboardnavigator(self):
        return self.dict_onboardnavigator

    def return_dict_city_data(self):
        return self.dict_city_data

    def return_dict_home_facts(self):
        return self.dict_home_facts

    def return_dict_offenders(self):
        return self.dict_offenders

    def return_dict_neighborhoodscout(self):
        return self.dict_neighborhoodscout

    def return_dict_bestplaces(self):
        return self.dict_bestplaces

    def return_dict_crime_total(self):
        return self.dict_crime_total

# copy all dictionaries to xls file
    def xls_new_sheet_create(self):
                wb = openpyxl.load_workbook(self.xls_name)
                if wb.sheetnames.count(self.full_addr[:25]) == 0:
                    example_sheet = wb["example"]
                    wb.copy_worksheet(example_sheet)
                    # print(wb.sheetnames)
                    new_sheet = wb['example Copy']
                    new_sheet.title = self.full_addr[:25]
                    # print(wb.sheetnames)
                    wb.save(self.xls_name)
                    print("XLS new sheet is ready, sheet name: {}".format(self.full_addr[:25]))
                    logging.debug("XLS new sheet is ready, sheet name: {}".format(self.full_addr[:25]))
                    wb.close()
                    return True
                else:
                    print("address is already exists in database!, recopy new run info ")
                    logging.debug("address is already exists in database!, recopy new run info ")
                    return False

    def all_dicts_to_xls(self):
        try:
            wb = openpyxl.load_workbook(self.xls_name)
            sheet = wb[self.full_addr[:25]]

            sheet['A3'].value = self.dict_basic_info['street']
            sheet['C3'].value = self.dict_basic_info['city']
            sheet['D3'].value = self.dict_basic_info['state']

            sheet['B24'].value = self.dict_onboardnavigator['Total personal']
            sheet['B25'].value = self.dict_onboardnavigator['Total property']
            sheet['B26'].value = self.dict_onboardnavigator['Total overall']
            sheet['B27'].value = self.dict_onboardnavigator['Year']

            sheet['B29'].value = self.dict_city_data['Crime Index city']
            sheet['B30'].value = self.dict_city_data['US avarage']
            sheet['B31'].value = self.dict_city_data['Pic of graph']
            sheet['B32'].value = self.dict_city_data['total info']

            sheet['B34'].value = self.dict_home_facts['Overall Score']
            sheet['B35'].value = self.dict_home_facts['Overall score big num']
            sheet['B36'].value = self.dict_home_facts['Score small procents']
            sheet['B37'].value = self.dict_offenders['offender1']
            sheet['B38'].value = self.dict_offenders['offender2']
            sheet['B39'].value = self.dict_offenders['offender3']
            sheet['B40'].value = self.dict_home_facts['Year']

            sheet['B42'].value = self.dict_neighborhoodscout['Diagram']
            sheet['B43'].value = self.dict_neighborhoodscout['List of safe areas']

            sheet['B45'].value = self.dict_bestplaces['Violent crime & US average']
            sheet['B46'].value = self.dict_bestplaces['Property crime & US average']
            sheet['B47'].value = self.dict_bestplaces['Photos and Maps of the city']

            wb.save(self.xls_name)
            wb.close()
            # printing the process
            print("Elements saved in {}".format(self.xls_name))
            logging.debug("Elements saved in {}".format(self.xls_name))
            return True
        except:
            return False



