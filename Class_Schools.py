'''
SCHOOLS Class was build for adding the crime information of the given address
Using address and collect information from all source web sites using Beautiful Soup and store it in dictionaries then in MySQL
'''
import openpyxl
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import time
import pprint
import logging
from selenium.webdriver.support.ui import Select


class Schools(object):
    def __init__(self, street, state, city, short_state, xls_name, county_name):
        # all setup params
        self.street = street
        self.state = state
        self.city = city
        self.short_state = short_state
        self.county = county_name
        self.xls_name = xls_name
        self.full_addr = self.street.lower() + " " + self.city.lower() + " " + self.state.lower() + " " + self.short_state.lower()
        self.driver = webdriver.Chrome("/Users/alexdezho/Downloads/chromedriver")

        #urls
        self.greatschools_url = 'https://www.greatschools.org/'
        self.schooldigger_url = 'https://www.schooldigger.com/'
        self.homefacts_url = 'https://www.homefacts.com/'
        self.niche_url = 'https://www.niche.com/?ref=k12'

        self.dict_schools_general = {
            'school - elementary name': 'NA',
            'school - elementary link': 'NA',
            'school - middle name': 'NA',
            'school - middle link': 'NA',
            'school - high name': 'NA',
            'school - high link': 'NA',
            'school - HF elementary name': 'NA',
            'school - HF elementary link': 'NA',
            'school - HF middle name': 'NA',
            'school - HF middle link': 'NA',
            'school - HF high name': 'NA',
            'school - HF high link': 'NA'

        }


        #dictionaries
        self.dict_basic_info = {
            'street': self.street,
            'city': self.city,
            'short_state': self.short_state,
            'state': self.state,
            'county': self.county
        }
        self.dict_greatschools = {
            'school - elementary name': 'NA',
            'school - elementary link': 'NA',
            'school - middle name': 'NA',
            'school - middle link': 'NA',
            'school - high name': 'NA',
            'school - high link': 'NA'

        }
        self.dict_schooldigger = {
            'school - elementary name': 'NA',
            'school - elementary link': 'NA',
            'school - middle name': 'NA',
            'school - middle link': 'NA',
            'school - high name': 'NA',
            'school - high link': 'NA'
        }
        self.dict_homefacts = {
            'school - elementary name': 'NA',
            'school - elementary link': 'NA',
            'school - middle name': 'NA',
            'school - middle link': 'NA',
            'school - high name': 'NA',
            'school - high link': 'NA'

        }
        self.dict_niche = {
            'link - County Schools': 'NA',
            'name - global': 'NA',
            'rank - School Districts if exists': 'NA',
            'grade - overall niche grade': 'NA',
            'link - all ranks state county schools/metropolitan/national': 'NA'

        }

    def closeBrowser(self):
        self.driver.close()
        logging.debug('Browser closed')
        print('Browser closed')

    def greateschools_to_dict(self):
            try:  # connecting to greateschools
                driver = self.driver
                driver.get(self.greatschools_url)
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="home-page"]/div[1]/div/section/div[1]/div[1]/div/div/div/div[1]/form/input')))
                driver.find_element_by_xpath('//*[@id="home-page"]/div[1]/div/section/div[1]/div[1]/div/div/div/div[1]/form/input').click()
                time.sleep(2)
                driver.find_element_by_xpath('//*[@id="home-page"]/div[1]/div/section/div[1]/div[1]/div/div/div/div[1]/form/input').send_keys(self.street.lower() + " " + self.city.lower() + " " + self.state.lower())
                time.sleep(2)
                driver.find_element_by_xpath('//*[@id="home-page"]/div[1]/div/section/div[1]/div[1]/div/div/div/div[2]/button/span[2]').click()
                driver.find_element_by_xpath('//*[@id="home-page"]/div[1]/div/section/div[1]/div[1]/div/div/div/div[2]/button/span[2]').click()
                time.sleep(5)
                print(driver.current_url)
                try:
                    # elementary school assigned tags
                    time.sleep(10)
                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, 'Elementary School')))
                    driver.find_element_by_partial_link_text('Elementary School').click()
                    time.sleep(3)
                    self.dict_greatschools['school - elementary link'] = driver.find_element_by_xpath('//*[@id="hero"]/div/div[2]/div[2]/div[1]/div/a/div[1]').text
                    print(self.dict_greatschools['school - elementary link'])
                    school_name = driver.find_element_by_xpath('//*[@id="hero"]/div/div[1]/h1').text
                    self.dict_greatschools['school - elementary name'] = school_name
                    print('Elementary school name: {}'.format(school_name))
                    self.dict_schools_general['school - elementary link'] = self.dict_greatschools['school - elementary link']
                    self.dict_schools_general['school - elementary name'] = self.dict_greatschools['school - elementary name']
                    driver.back()
                except:
                    print('failed to locate elemantary school from greateschools')

                try:
                    # middle school assigned tags
                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, 'Middle School')))
                    driver.find_element_by_partial_link_text('Middle School').click()
                    time.sleep(3)
                    self.dict_greatschools['school - middle link'] = driver.find_element_by_xpath('//*[@id="hero"]/div/div[2]/div[2]/div[1]/div/a/div[1]').text
                    print(self.dict_greatschools['school - middle link'])
                    school_name = driver.find_element_by_xpath('//*[@id="hero"]/div/div[1]/h1').text
                    print('Middle school name: {}'.format(school_name))
                    self.dict_greatschools['school - middle name'] = school_name
                    self.dict_schools_general['school - middle link'] = self.dict_greatschools['school - middle link']
                    self.dict_schools_general['school - middle name'] = self.dict_greatschools['school - middle name']
                    driver.back()
                except:
                    print('failed to locate middle school from greateschools')

                try:
                    # high school assigned tags
                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, 'High School')))
                    driver.find_element_by_partial_link_text('High School').click()
                    time.sleep(3)
                    self.dict_greatschools['school - high link'] = driver.find_element_by_xpath('//*[@id="hero"]/div/div[2]/div[2]/div[1]/div/a/div[1]').text
                    print(self.dict_greatschools['school - high link'])
                    school_name = driver.find_element_by_xpath('//*[@id="hero"]/div/div[1]/h1').text
                    print('High school name: {}'.format(school_name))
                    self.dict_greatschools['school - high name'] = school_name
                    self.dict_schools_general['school - high link'] = self.dict_greatschools['school - high link']
                    self.dict_schools_general['school - high name'] = self.dict_greatschools['school - high name']
                    driver.back()
                except:
                    print('failed to locate high school from greateschools')
            except:
                print('something went wrong with greateschools')

    def schooldigger_to_dict(self): #check
        try:
            driver = self.driver
            driver.get(self.schooldigger_url)
            print(driver.current_url)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="txtHPAC"]')))
            time.sleep(2)
            driver.find_element_by_xpath('//*[@id="txtHPAC"]').click()
            time.sleep(1)
            driver.find_element_by_xpath('//*[@id="txtHPAC"]').send_keys(self.street.lower() + " " + self.city.lower() + " " + self.state.lower())
            time.sleep(1)
            driver.find_element_by_xpath('//*[@id="txtHPAC"]').send_keys(Keys.ENTER)
            time.sleep(3)
            print(driver.current_url)

            # elementary schools under boundary tags
            try:
                driver.find_element_by_partial_link_text('Elementary').click()
                time.sleep(2)
                WebDriverWait(driver, 10).until(EC.presence_of_element_located(
                    (By.XPATH, '//*[@id="aspnetForm"]/div[5]/div[1]/div[3]/h1/span')))
                self.dict_schooldigger['school - elementary link'] = driver.current_url
                self.dict_schooldigger['school - elementary name'] = driver.find_element_by_xpath('//*[@id="aspnetForm"]/div[5]/div[1]/div[3]/h1/span').text
                print('elemantary school found in schooldigger')
                driver.back()
                time.sleep(2)
            except:
                print('elemantary school was not fount in schooldigger')
            # middle boundry tags

            try:
                driver.find_element_by_partial_link_text('Middle').click()
                time.sleep(2)
                WebDriverWait(driver, 10).until(EC.presence_of_element_located(
                    (By.XPATH, '//*[@id="aspnetForm"]/div[5]/div[1]/div[3]/h1/span')))
                self.dict_schooldigger['school - middle link'] = driver.current_url
                self.dict_schooldigger['school - middle name'] = driver.find_element_by_xpath('//*[@id="aspnetForm"]/div[5]/div[1]/div[3]/h1/span').text
                driver.back()
                time.sleep(2)
                #print('middle school found in schooldigger')
            except:
                print('middle school was not fount in schooldigger')
            # high boundary tags
            try:
                driver.find_element_by_partial_link_text('High').click()
                time.sleep(2)
                WebDriverWait(driver, 10).until(EC.presence_of_element_located(
                    (By.XPATH, '//*[@id="aspnetForm"]/div[5]/div[1]/div[3]/h1/span')))
                self.dict_schooldigger['school - high link'] = driver.current_url
                self.dict_schooldigger['school - high name'] = driver.find_element_by_xpath('//*[@id="aspnetForm"]/div[5]/div[1]/div[3]/h1/span').text
                #print('high school found in schooldigger')
                driver.back()
                time.sleep(1)
            except:
                print('high school was not fount in schooldigger')

            print('schooldigger params was copied to dictionary , success {}'.format(self.dict_schooldigger))
            return True
        except:
            print('failed to connect or locate params from schooldigger')
            logging.debug('fail')
            return False

    def homefacts_to_dict(self):
        try:
            driver = self.driver
            driver.get(self.homefacts_url)
            time.sleep(2)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="fulladdress"]')))
            driver.find_element_by_xpath('//*[@id="fulladdress"]').click()
            time.sleep(1)
            driver.find_element_by_xpath('//*[@id="fulladdress"]').send_keys(self.street.lower() + " " + self.city.lower() + " " + self.state.lower())
            time.sleep(1)
            driver.find_element_by_xpath('//*[@id="fulladdress"]').send_keys(Keys.ENTER)
            time.sleep(3)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="navbar"]/ul/li[4]/a')))
            driver.find_element_by_xpath('//*[@id="navbar"]/ul/li[4]/a').click()
            time.sleep(2)
            driver.execute_script("window.scrollTo(0,550)")
            time.sleep(3)
            print(driver.current_url)
            time.sleep(3)
            # elementary
            try:
                driver.execute_script("window.scrollTo(0,750)")
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, 'ELEMENTARY SCHOOL')))
                driver.find_element_by_partial_link_text('ELEMENTARY SCHOOL').click()
                time.sleep(5)
                school_name = driver.find_element_by_xpath('/html/body/section[2]/div[2]/div/div[1]/h1/span').text
                print(school_name)
                time.sleep(3)
                self.dict_homefacts['school - elementary name'] = school_name #                 //*[@id="school_year_2019"]/div[1]/div[2]
                self.dict_homefacts['school - elementary link'] = driver.find_element_by_xpath('//*[@id="school_year_2018"]/div[1]/div[2]').get_attribute('class')
                print(self.dict_homefacts['school - elementary link'])
                time.sleep(2)
                self.dict_schools_general['school - HF elementary name'] = self.dict_homefacts['school - elementary name']
                self.dict_schools_general['school - HF elementary link'] = self.dict_homefacts['school - elementary link']
                print('elemantary school found in homefacts')
                driver.back()
                time.sleep(5)
            except:
                print('elemantary school was not fount in homefacts')
                #driver.back()
            # middle
            try:
                driver.execute_script("window.scrollTo(0,750)")
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, 'MIDDLE SCHOOL')))
                driver.find_element_by_partial_link_text('MIDDLE SCHOOL').click()
                time.sleep(5)
                school_name = driver.find_element_by_xpath('/html/body/section[2]/div[2]/div/div[1]/h1/span').text
                print(school_name)
                self.dict_homefacts['school - middle name'] = school_name
                self.dict_homefacts['school - middle link'] = driver.find_element_by_xpath('//*[@id="school_year_2018"]/div[1]/div[2]').get_attribute('class')
                print(self.dict_homefacts['school - middle link'])
                time.sleep(2)
                self.dict_schools_general['school - HF middle name'] = self.dict_homefacts['school - middle name']
                self.dict_schools_general['school - HF middle link'] = self.dict_homefacts['school - middle link']
                driver.back()
                time.sleep(5)
            except:
                print('middle school was not fount in homefacts')
                #driver.back()

            # high
            try:
                print('trying to locate high school')
                driver.execute_script("window.scrollTo(0,750)")
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, 'HIGH SCHOOL')))
                driver.find_element_by_partial_link_text('HIGH').click()
                time.sleep(5)
                school_name = driver.find_element_by_xpath('/html/body/section[2]/div[2]/div/div[1]/h1/span').text
                print(school_name)
                time.sleep(5)
                self.dict_homefacts['school - high name'] = school_name
                print('trying to locate high school grade from pic') # //*[@id="school_year_2018"]/div[1]/div[2]
                self.dict_homefacts['school - high link'] = driver.find_element_by_xpath('//*[@id="school_year_2018"]/div[1]/div[2]').get_attribute('class')
                print(self.dict_homefacts['school - high link'])
                time.sleep(2)
                self.dict_schools_general['school - HF high name'] = self.dict_homefacts['school - high name']
                self.dict_schools_general['school - HF high link'] = self.dict_homefacts['school - high link']
                driver.back()
                time.sleep(5)
            except:
                print('high school was not fount in homefacts')
                #driver.back()

            print('homefacts params was copied to dictionary , success {}'.format(self.dict_homefacts))

        except:
            print('fail to copy params from homefacts')
            logging.debug('fail')

    def niche_to_dict(self):
        try:
            driver = self.driver
            driver.get(self.niche_url)
            time.sleep(4)
            driver.find_element_by_xpath(
                '//*[@id="maincontent"]/div/section[1]/section/div/ul/li[2]/div[1]/input').click()
            driver.find_element_by_xpath(
                '//*[@id="maincontent"]/div/section[1]/section/div/ul/li[2]/div[1]/input').send_keys(
                self.county + ' county')
            time.sleep(1)
            driver.find_element_by_xpath(
                '//*[@id="maincontent"]/div/section[1]/section/div/ul/li[2]/div[1]/input').sendKeys(Keys.ENTER)
            time.sleep(4)
            self.dict_niche['link - County Schools'] = driver.current_url
            self.dict_niche['name - global'] = driver.current_url
            self.dict_niche['rank - School Districts if exists'] = driver.current_url
            self.dict_niche['grade - overall niche grade'] = driver.current_url
            self.dict_niche['link - all ranks state county schools/metropolitan/national'] = driver.current_url

            print('niche params was copied to dictionary , success {} '.format(self.dict_niche))
            return True
        except:
            print('fail to locate params from niche')
            logging.debug('fail')
            return False

    def printall(self):
        pp = pprint.PrettyPrinter(indent=4)
        print('greate schools')
        pp.pprint(self.dict_greatschools)
        print('school digger')
        pp.pprint(self.dict_schooldigger)
        print('home facts')
        pp.pprint(self.dict_homefacts)
        print('niche')
        pp.pprint(self.dict_niche)

# returning all dictionaries for future use to add to general list
    def return_dict_basic_info(self):
        return self.dict_basic_info
    def return_dict_greateshcools(self):
        return self.dict_greatschools
    def return_dict_schooldigger(self):
        return self.dict_schooldigger
    def return_dict_homefacts(self):
        return self.dict_homefacts
    def return_dict_niche(self):
        return self.dict_niche
    def return_dict_schools_general(self):
        return self.dict_schools_general
# copy all dictionaries to xls file
    def xls_new_sheet_for_search_create(self):
        wb = openpyxl.load_workbook(self.xls_name)
        if wb.sheetnames.count(self.full_addr[:25]) == 0:
            example_sheet = wb["example"]
            wb.copy_worksheet(example_sheet)
            # print(wb.sheetnames)
            new_sheet = wb['example Copy']
            new_sheet.title = self.full_addr[:25]
            # print(wb.sheetnames)
            wb.save(self.xls_name)
            print("XLS new sheet name: {}".format(self.full_addr[:25]))
            logging.debug("XLS new sheet is ready, sheet name: {}".format(self.full_addr[:25]))
            wb.close()
            return True
        else:
            print("address was already searched & exists in database recopy new params")
            logging.debug("address was already searched & exists in database")
            return False
    def all_dicts_to_xls(self):
        wb = openpyxl.load_workbook(self.xls_name)
        sheet = wb[self.full_addr[:25]]
        # print(wb.sheetnames)
        sheet['B2'].value = self.dict_greatschools['school - elementary name']
        sheet['B3'].value = self.dict_greatschools['school - elementary link']
        sheet['B4'].value = self.dict_greatschools['school - middle name']
        sheet['B5'].value = self.dict_greatschools['school - middle link']
        sheet['B6'].value = self.dict_greatschools['school - high name']
        sheet['B7'].value = self.dict_greatschools['school - high link']

        sheet['B9'].value = self.dict_schooldigger['school - elementary name']
        sheet['B10'].value = self.dict_schooldigger['school - elementary link']
        sheet['B11'].value = self.dict_schooldigger['school - middle name']
        sheet['B12'].value = self.dict_schooldigger['school - middle link']
        sheet['B13'].value = self.dict_schooldigger['school - high name']
        sheet['B14'].value = self.dict_schooldigger['school - high link']

        sheet['B16'].value = self.dict_homefacts['school - elementary name']
        sheet['B17'].value = self.dict_homefacts['school - elementary link']
        sheet['B18'].value = self.dict_homefacts['school - middle name']
        sheet['B19'].value = self.dict_homefacts['school - middle link']
        sheet['B20'].value = self.dict_homefacts['school - high name']
        sheet['B21'].value = self.dict_homefacts['school - high link']

        #sheet['B23'].value = self.dict_niche['link - County Schools']
        #sheet['B24'].value = self.dict_niche['name - global']
        #sheet['B25'].value = self.dict_niche['rank - School Districts if exists']
        #sheet['B26'].value = self.dict_niche['grade - overall niche grade']
        #sheet['B27'].value = self.dict_niche['link - all ranks state county schools/metropolitan/national']

        sheet['B29'].value = self.dict_basic_info['street']
        sheet['B30'].value = self.dict_basic_info['city']
        sheet['B31'].value = self.dict_basic_info['state']

        wb.save(self.xls_name)
        wb.close()
        # printing the process
        print("Dictionaries was completed & saved in {}".format(self.xls_name))
        logging.debug("Dictionaries was completed & saved in {}".format(self.xls_name))
        return True

