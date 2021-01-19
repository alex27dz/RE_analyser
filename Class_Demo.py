
import openpyxl
# import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import time
import pprint
import logging
from selenium.webdriver.support.ui import Select
import mysql.connector
import datetime
from selenium.webdriver.common.action_chains import ActionChains


class HometownLocator(object):
    def __init__(self, street, state, city, short_state, xls_name):
        self.driver = webdriver.Chrome("/Users/alexdezho/Downloads/chromedriver")
        self.stateformetro = state
        # websites operation
        self.googleMaps_url = "https://www.google.com/maps/"
        self.coordinate_search_url = "https://" + state.replace(" ", "").lower() + ".hometownlocator.com/maps"
        self.metropol_Tool_url = "https://www.huduser.gov/portal/datasets/geotool/select_Geography.odn"

        self.coordinates_url_track = ''
        # hometownlocator urls
        self.census_block_url = ''
        self.census_track_url = ''
        self.zip_code = ''
        self.city_url = ''
        self.metropolitan_url = "https://" + str(state).replace(" ", "").lower() + ".hometownlocator.com/cities/msa/"

        # all setup params
        self.street = street
        self.state = state
        self.city = city
        self.short_state = short_state
        self.xls_name = xls_name
        self.full_addr = self.street.lower() + " " + self.city.lower() + " " + self.state.replace(" ", "").lower() + " " + self.short_state.lower()

        # into functions operational parameters

        self.zip_code_url = ' '
        self.coord = " "
        self.county = " "
        self.county_url = " "
        self.state_search = " "
        self.county_search = " "
        self.metropolitan_name = " "
        self.metropolitan_url_htl = ' '
        self.index = " "
        self.google_maps_link = ' '

        # dictionaries
        self.dict_basic_info = {
            'street': self.street,
            'city': self.city,
            'short_state': self.short_state,
            'county': '',
            'zip_code': '',
            'metropolitan': '',
            'link_google_maps': '',
            'coordinates': ''
        }
        self.dict_block = {
            "Total_Population": "",
            "Population_Growth_2010_2019": "",
            "Population_Growth_2019_2024": "",
            "Median_Household_Income": "",
            "Average_Household_Income": "",
            "Owner_Occupied_HU": "",
            "Renter_Occupied_HU": "",
            "Vacant_Housing_Units": "",
            "Median_Home_Value": "",
            "Total_Hoouseholds": "",
            "Avarage_Households_Size": "",
            "Family_Households": ""
        }
        self.dict_track = {
            "Total_Population": "",
            "Population_Growth_2010_2019": "",
            "Population_Growth_2019_2024": "",
            "Median_Household_Income": "",
            "Average_Household_Income": "",
            "Owner_Occupied_HU": "",
            "Renter_Occupied_HU": "",
            "Vacant_Housing_Units": "",
            "Median_Home_Value": "",
            "Total_Hoouseholds": "",
            "Avarage_Households_Size": "",
            "Family_Households": ""
        }
        self.dict_zip_code = {
            "Total_Population": "",
            "Population_Growth_2010_2019": "",
            "Population_Growth_2019_2024": "",
            "Median_Household_Income": "",
            "Average_Household_Income": "",
            "Owner_Occupied_HU": "",
            "Renter_Occupied_HU": "",
            "Vacant_Housing_Units": "",
            "Median_Home_Value": "",
            "Total_Hoouseholds": "",
            "Avarage_Households_Size": "",
            "Family_Households": ""
        }
        self.dict_city = {
            "Total_Population": "",
            "Population_Growth_2010_2019": "",
            "Population_Growth_2019_2024": "",
            "Median_Household_Income": "",
            "Average_Household_Income": "",
            "Owner_Occupied_HU": "",
            "Renter_Occupied_HU": "",
            "Vacant_Housing_Units": "",
            "Median_Home_Value": "",
            "Total_Hoouseholds": "",
            "Avarage_Households_Size": "",
            "Family_Households": ""
        }
        self.dict_county = {
            "Total_Population": "",
            "Population_Growth_2010_2019": "",
            "Population_Growth_2019_2024": "",
            "Median_Household_Income": "",
            "Average_Household_Income": "",
            "Owner_Occupied_HU": "",
            "Renter_Occupied_HU": "",
            "Vacant_Housing_Units": "",
            "Median_Home_Value": "",
            "Total_Hoouseholds": "",
            "Avarage_Households_Size": "",
            "Family_Households": ""
        }
        self.dict_metro = {
            "Total_Population": "",
            "Population_Growth_2010_2019": "",
            "Population_Growth_2019_2024": "",
            "Median_Household_Income": "",
            "Average_Household_Income": "",
            "Owner_Occupied_HU": "",
            "Renter_Occupied_HU": "",
            "Vacant_Housing_Units": "",
            "Median_Home_Value": "",
            "Total_Hoouseholds": "",
            "Avarage_Households_Size": "",
            "Family_Households": ""
        }

    def closeBrowser(self):
        self.driver.close()
        logging.debug('Browser closed')
        print('Browser closed')

    def google_Maps_Addr_Coord(self):
        try:
            driver = self.driver
            driver.get(self.googleMaps_url)
            # await command waiting till we find the element then continue
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="searchboxinput"]')))
            time.sleep(10)
            driver.find_element_by_xpath('//*[@id="searchboxinput"]').send_keys(self.full_addr)
            time.sleep(1)
            driver.find_element_by_xpath('//*[@id="searchbox-searchbutton"]').click()
            time.sleep(3)
            driver.find_element_by_xpath('//*[@id="widget-zoom-in"]').click()  # zoom1
            time.sleep(1)
            driver.find_element_by_xpath('//*[@id="widget-zoom-in"]').click()  # zoom2
            time.sleep(1)
            driver.find_element_by_xpath('//*[@id="widget-zoom-in"]').click()  # zoom3
            time.sleep(1)
            driver.find_element_by_xpath('//*[@id="widget-zoom-in"]').click()  # zoom4
            time.sleep(3)
            self.coord = self.driver.current_url
            index = self.coord.find("@") + 1
            self.coord = self.coord[index:index + 21]
            self.coord = self.coord.replace(",", ":")
            self.dict_basic_info['coordinates'] = self.coord
            self.google_maps_link = driver.current_url
            self.dict_basic_info['link_google_maps'] = self.google_maps_link
            print("google coordinates is: {}".format(self.coord))
            print('using url {}'.format(self.coordinate_search_url))
            driver.get(self.coordinate_search_url)  # fix and change coordinates link
            time.sleep(20)
            print('1 - scrolling to enter address')
            ActionChains(driver).move_to_element(driver.find_element_by_xpath('//*[@id="gcForm"]/fieldset/p/input[1]')).perform()
            time.sleep(5)
            driver.execute_script("window.scrollTo(0,450)")
            time.sleep(2)
            driver.find_element_by_xpath('//*[@id="gcForm"]/fieldset/p/input[1]').click()
            time.sleep(2)
            driver.find_element_by_xpath('//*[@id="gcForm"]/fieldset/p/input[1]').send_keys(self.street + ' ' + self.city + ' ' + self.short_state)
            time.sleep(2)
            driver.find_element_by_xpath('//*[@id="gcForm"]/fieldset/p/input[2]').click()
            time.sleep(10)
            print('preparing to locate general info')
            print('starting with trying to locate zip code')
            self.zip_code_url = driver.find_element_by_partial_link_text('ZIP Code ').text
            index1 = self.zip_code_url.find('Code') + 5
            self.zip_code_url = self.zip_code_url[index1:]
            self.dict_basic_info['zip_code'] = self.zip_code_url
            print('zip code located')
            print(self.dict_basic_info['zip_code'])
            self.zip_code_url = 'https://' + self.state.replace(" ", "").lower() + '.hometownlocator.com/zip-codes/data,zipcode,' + self.zip_code_url + '.cfm'
            print('trying to locate county url')
            self.county = driver.find_element_by_partial_link_text('County').text
            self.dict_basic_info['county'] = self.county
            self.county_url = "https://" + self.state.replace(" ", "").lower() + ".hometownlocator.com/" + self.short_state.lower() + "/" + str(self.county)[:-7].lower() + "/"
            print('county is {}'.format(self.county))
            print('county url located')
            print(self.county_url)
            print('trying to locate city url')
            self.city_url = "https://" + self.state.replace(" ", "").lower() + ".hometownlocator.com/" + self.short_state.lower() + "/" + str(self.county)[:-7].lower() + "/" + self.city.lower() + ".cfm"
            print('city url detected')
            print(self.city_url)
        except:
            print('failed - to locate all basic information')
            return False

    def metropolitan_area_Look_Up_Tool(self):
        # getting metropolitan name by using state name and county name
        self.state_search = self.state + " - " + self.short_state.upper()
        self.county_search = self.county + ", " + self.short_state.upper()
        try:
            driver = self.driver
            driver.get(self.metropol_Tool_url)
            time.sleep(10)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="select_geo_options"]/form[1]/div[1]/select')))
            # select state - table 1
            driver.find_element_by_xpath('//*[@id="select_geo_options"]/form[1]/div[1]/select')
            Select(driver.find_element_by_tag_name('select')).select_by_visible_text(self.state_search)
            time.sleep(1)
            logging.debug('first table state selected,success')
            # select - table 2
            driver.find_element_by_xpath('//*[@id="select_geo_options"]/form[1]/div[2]/select')
            Select(driver.find_element_by_id("countyselect")).select_by_visible_text(self.county_search)
            driver.find_element_by_xpath('//*[@id="select_geo_options"]/form[1]/input').click()
            # locate metropolitan name
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/p[2]/em')))
            self.metropolitan_name = driver.find_element_by_xpath('/html/body/p[2]/em').text
            self.metropolitan_name = str(self.metropolitan_name)
            self.index = self.metropolitan_name.find('-')
            self.metropolitan_name = self.metropolitan_name[:self.index]
            self.dict_basic_info['metropolitan'] = self.metropolitan_name
            logging.debug('Metropolitan name: {}'.format(self.metropolitan_name))
            print('Metropolitan found name: {}'.format(self.metropolitan_name))
        except:
            print('Metropolitan name Failed to found')
            return False

    def metro_to_url(self):
        try:
            driver = self.driver
            driver.get(self.metropolitan_url)
            time.sleep(1)
            driver.execute_script("window.scrollTo(0,615)")
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, self.metropolitan_name)))
            driver.find_element_by_partial_link_text(self.metropolitan_name).click()
            time.sleep(1)
            self.metropolitan_url_htl = driver.current_url
            #print('metropolitan url located!')
            logging.debug("Metro url: {}".format(self.metropolitan_url_htl))
            print("Metro url: {}".format(self.metropolitan_url_htl))
            return self.metropolitan_url_htl
        except:
            print('failed to locate metro url')
            return False

    def return_block_url(self):
        try:
            driver = self.driver
            print('trying to locate block url')
            print(self.coordinate_search_url)
            driver.get(self.coordinate_search_url)  # fix and change coordinates link
            time.sleep(20)
            print('scrolling to search address')
            ActionChains(driver).move_to_element(driver.find_element_by_xpath('//*[@id="gcForm"]/fieldset/p/input[1]')).perform()
            time.sleep(5)
            driver.execute_script("window.scrollTo(0,450)")
            time.sleep(2)
            driver.find_element_by_xpath('//*[@id="gcForm"]/fieldset/p/input[1]').click()
            time.sleep(2)
            driver.find_element_by_xpath('//*[@id="gcForm"]/fieldset/p/input[1]').send_keys(self.street + ' ' + self.city + ' ' + self.short_state)
            time.sleep(2)
            driver.find_element_by_xpath('//*[@id="gcForm"]/fieldset/p/input[2]').click()
            time.sleep(10)
            driver.execute_script("window.scrollTo(0,650)")
            time.sleep(3)
            self.census_block_url = driver.find_element_by_partial_link_text('(Census Block Group)').click()
            time.sleep(5)
            self.census_block_url = driver.current_url
            print(driver.current_url)
        except:
            print('failed to locate block url - not from 404 error')

        return self.census_block_url

    def return_track_url(self):
        try:
            driver = self.driver
            print('trying to locate track url')
            driver.get(self.coordinate_search_url)  # fix and change coordinates link
            time.sleep(20)
            print('1')
            ActionChains(driver).move_to_element(driver.find_element_by_xpath('//*[@id="gcForm"]/fieldset/p/input[1]')).perform()
            time.sleep(5)
            driver.execute_script("window.scrollTo(0,450)")
            time.sleep(2)
            driver.find_element_by_xpath('//*[@id="gcForm"]/fieldset/p/input[1]').click()
            time.sleep(2)
            driver.find_element_by_xpath('//*[@id="gcForm"]/fieldset/p/input[1]').send_keys(self.street + ' ' + self.city + ' ' + self.short_state)
            time.sleep(2)
            driver.find_element_by_xpath('//*[@id="gcForm"]/fieldset/p/input[2]').click()
            time.sleep(10)
            driver.execute_script("window.scrollTo(0,650)")
            time.sleep(3)
            self.census_track_url = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[3]/ul/li[2]/a').click()
            time.sleep(5)
            self.census_track_url = driver.current_url
            a = ""
            try:
                a = driver.find_element_by_xpath('//*[@id="content"]/div/fieldset/h2').text
            except:
                print('url should be founded')

            if a == '404 - File or directory not found.':
                print('failed to locate track url')
                self.dict_track["Total_Population"] = 'NA'
                self.dict_track["Population_Growth_2010_2019"] = 'NA'
                self.dict_track["Population_Growth_2019_2024"] = 'NA'
                self.dict_track["Median_Household_Income"] = 'NA'
                self.dict_track["Average_Household_Income"] = 'NA'
                self.dict_track["Total_Housing_Units"] = 'NA'
                self.dict_track["Owner_Occupied_HU"] = 'NA'
                self.dict_track["Renter_Occupied_HU"] = 'NA'
                self.dict_track["Vacant_Housing_Units"] = 'NA'
                self.dict_track["Median_Home_Value"] = 'NA'
            else:
                print('track url located')
                print(self.census_track_url)
        except:
            print('failed to locate track url not from 404 error')
        return self.census_track_url

    def return_zip_code_url(self):
        try:
            driver = self.driver
            print('trying to locate zip code url')
            #print(self.coordinate_search_url)
            driver.get(self.coordinate_search_url)
            time.sleep(3)
            driver.execute_script("window.scrollTo(0,615)")
            time.sleep(3)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="bodycontainer"]/div[3]/div[1]/div[3]/ul/li[4]/a')))
            self.zip_code_url = driver.find_element_by_partial_link_text('ZIP Code ').click()
            time.sleep(3)
            self.zip_code_url = driver.current_url
            a = ""
            try:
                a = driver.find_element_by_xpath('//*[@id="content"]/div/fieldset/h2').text
            except:
                print('url should be founded')

            if a == '404 - File or directory not found.':
                print('failed to locate zip code url')
                self.dict_zip_code["Total_Population"] = 'NA'
                self.dict_zip_code["Population_Growth_2010_2019"] = 'NA'
                self.dict_zip_code["Population_Growth_2019_2024"] = 'NA'
                self.dict_zip_code["Median_Household_Income"] = 'NA'
                self.dict_zip_code["Average_Household_Income"] = 'NA'
                self.dict_zip_code["Total_Housing_Units"] = 'NA'
                self.dict_zip_code["Owner_Occupied_HU"] = 'NA'
                self.dict_zip_code["Renter_Occupied_HU"] = 'NA'
                self.dict_zip_code["Vacant_Housing_Units"] = 'NA'
                self.dict_zip_code["Median_Home_Value"] = 'NA'
            else:
                print('zip code url located')
                print(self.zip_code_url)

        except:

            print('failed to locate zip code url not from 404 error')

        return self.zip_code_url

    def return_county_url(self):
        try:
            print('trying to locate county url')
            driver = self.driver
            driver.get(self.coordinate_search_url)
            time.sleep(20)
            print('1')
            ActionChains(driver).move_to_element(driver.find_element_by_xpath('//*[@id="gcForm"]/fieldset/p/input[1]')).perform()
            time.sleep(5)
            driver.execute_script("window.scrollTo(0,450)")
            time.sleep(2)
            driver.find_element_by_xpath('//*[@id="gcForm"]/fieldset/p/input[1]').click()
            time.sleep(2)
            driver.find_element_by_xpath('//*[@id="gcForm"]/fieldset/p/input[1]').send_keys(self.street + ' ' + self.city + ' ' + self.short_state)
            time.sleep(2)
            driver.find_element_by_xpath('//*[@id="gcForm"]/fieldset/p/input[2]').click()
            time.sleep(10)
            driver.execute_script("window.scrollTo(0,650)")
            time.sleep(5)
            driver.find_element_by_partial_link_text('County').click()
            time.sleep(5)
            self.county_url = driver.current_url
            return self.county_url
        except:
            print('could not return county url')

    def return_city_url(self):
        return self.city_url

    def return_metro_url(self):
        return self.metropolitan_url_htl

    def HTML_to_dictionary(self, url):
        dict = {
            "Total_Population": "",
            "Population_Growth_2010_2019": "",
            "Population_Growth_2019_2024": "",
            "Median_Household_Income": "",
            "Average_Household_Income": "",
            "Owner_Occupied_HU": "",
            "Renter_Occupied_HU": "",
            "Vacant_Housing_Units": "",
            "Median_Home_Value": "",
            "Total_Hoouseholds": "",
            "Avarage_Households_Size": "",
            "Family_Households": ""
        }
        try:
            driver = self.driver
            driver.get(url)
            print(url)
            print('trying to locate elements path 1')
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="bodycontainer"]/div[4]/div[1]/div[5]/table/tbody/tr[2]/td[2]')))
            print('printing for checking the first element')
            print(driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[5]/table/tbody/tr[2]/td[2]').text)
            time.sleep(2)
            print('1')
            dict["Total_Population"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[5]/table/tbody/tr[2]/td[2]').text
            print('2')
            dict["Population_Growth_2010_2019"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[9]/table/tbody/tr[2]/td[2]').text
            print('3')
            dict["Population_Growth_2019_2024"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[9]/table/tbody/tr[2]/td[3]').text
            print('4')
            dict["Median_Household_Income"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[5]/table/tbody/tr[11]/td[2]').text
            print('5')
            dict["Average_Household_Income"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[5]/table/tbody/tr[12]/td[2]').text
            print('6')
            dict["Total_Housing_Units"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[6]/table/tbody/tr[2]/td[2]').text
            print('7')
            dict["Owner_Occupied_HU"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[6]/table/tbody/tr[3]/td[2]').text
            print('8')
            dict["Renter_Occupied_HU"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[6]/table/tbody/tr[4]/td[2]').text
            print('9')
            dict["Vacant_Housing_Units"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[6]/table/tbody/tr[5]/td[2]').text
            print('10')
            dict["Median_Home_Value"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[6]/table/tbody/tr[6]/td[2]').text
            print('11')
            dict["Total_Hoouseholds"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[6]/table/tbody/tr[11]/td[2]').text
            print('12')
            dict["Avarage_Households_Size"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[6]/table/tbody/tr[12]/td[2]').text
            print('13')
            dict["Family_Households"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[6]/table/tbody/tr[13]/td[2]').text
            print('14')
            print('HTML page params was copied to dict success------------------------------')
        except:
            try:
                driver = self.driver
                print('trying to locate elements with another path 2')
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="bodycontainer"]/div[4]/div[1]/div[3]/table/tbody/tr[2]/td[2]')))
                print('printing for checking the first element')
                print(driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[3]/table/tbody/tr[2]/td[2]').text)
                time.sleep(2)
                print('1')
                dict["Total_Population"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[3]/table/tbody/tr[2]/td[2]').text
                print('2')
                dict["Population_Growth_2010_2019"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[7]/table/tbody/tr[2]/td[2]').text
                print('3')
                dict["Population_Growth_2019_2024"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[7]/table/tbody/tr[2]/td[3]').text
                print('4')
                dict["Median_Household_Income"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[3]/table/tbody/tr[11]/td[2]').text
                print('5')
                dict["Average_Household_Income"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[3]/table/tbody/tr[12]/td[2]').text
                print('6')
                dict["Total_Housing_Units"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[4]/table/tbody/tr[2]/td[2]').text
                print('7')
                dict["Owner_Occupied_HU"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[4]/table/tbody/tr[3]/td[2]').text
                print('8')
                dict["Renter_Occupied_HU"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[4]/table/tbody/tr[4]/td[2]').text
                print('9')
                dict["Vacant_Housing_Units"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[4]/table/tbody/tr[5]/td[2]').text
                print('10')
                dict["Median_Home_Value"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[4]/table/tbody/tr[6]/td[2]').text
                print('11')
                dict["Total_Hoouseholds"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[4]/table/tbody/tr[11]/td[2]').text
                print('12')
                dict["Avarage_Households_Size"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[4]/table/tbody/tr[12]/td[2]').text
                print('13')
                dict["Family_Households"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[1]/div[4]/table/tbody/tr[13]/td[2]').text
                print('14')
                print('HTML page params was copied to dict success------------------------------')
            except:
                try:
                    driver = self.driver
                    print('trying to locate elements with another path 3')
                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="bodycontainer"]/div[3]/div[1]/div[4]/table/tbody/tr[2]/td[2]')))
                    print('printing for checking the first element')
                    print(driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[4]/table/tbody/tr[2]/td[2]').text)
                    time.sleep(2)
                    print('1')
                    dict["Total_Population"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[4]/table/tbody/tr[2]/td[2]').text
                    print('2')
                    dict["Population_Growth_2010_2019"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[8]/table/tbody/tr[2]/td[2]').text
                    print('3')
                    dict["Population_Growth_2019_2024"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[8]/table/tbody/tr[2]/td[3]').text
                    print('4')
                    dict["Median_Household_Income"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[4]/table/tbody/tr[11]/td[2]').text
                    print('5')
                    dict["Average_Household_Income"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[4]/table/tbody/tr[12]/td[2]').text
                    print('6')
                    dict["Total_Housing_Units"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[5]/table/tbody/tr[2]/td[2]').text
                    print('7')
                    dict["Owner_Occupied_HU"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[5]/table/tbody/tr[3]/td[2]').text
                    print('8')
                    dict["Renter_Occupied_HU"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[5]/table/tbody/tr[4]/td[2]').text
                    print('9')
                    dict["Vacant_Housing_Units"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[5]/table/tbody/tr[5]/td[2]').text
                    print('10')
                    dict["Median_Home_Value"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[5]/table/tbody/tr[6]/td[2]').text
                    print('11')
                    dict["Total_Hoouseholds"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[5]/table/tbody/tr[11]/td[2]').text
                    print('12')
                    dict["Avarage_Households_Size"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[5]/table/tbody/tr[12]/td[2]').text
                    print('13')
                    dict["Family_Households"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[5]/table/tbody/tr[13]/td[2]').text
                    print('14')
                    print('HTML page params was copied to dict success------------------------------')
                except:
                    try:
                        driver = self.driver
                        print('trying to locate elements with another third path 4')
                        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="bodycontainer"]/div[3]/div[1]/div[6]/table/tbody/tr[2]/td[2]')))
                        print('printing for checking the first element')
                        print(driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[6]/table/tbody/tr[2]/td[2]').text)
                        time.sleep(2)
                        print('1')
                        dict["Total_Population"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[6]/table/tbody/tr[2]/td[2]').text
                        print('2')
                        dict["Population_Growth_2010_2019"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[10]/table/tbody/tr[2]/td[2]').text
                        print('3')
                        dict["Population_Growth_2019_2024"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[10]/table/tbody/tr[2]/td[3]').text
                        print('4')
                        dict["Median_Household_Income"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[6]/table/tbody/tr[11]/td[2]').text
                        print('5')
                        dict["Average_Household_Income"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[6]/table/tbody/tr[12]/td[2]').text
                        print('6')
                        dict["Total_Housing_Units"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[7]/table/tbody/tr[2]/td[2]').text
                        print('7')
                        dict["Owner_Occupied_HU"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[7]/table/tbody/tr[3]/td[2]').text
                        print('8')
                        dict["Renter_Occupied_HU"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[7]/table/tbody/tr[4]/td[2]').text
                        print('9')
                        dict["Vacant_Housing_Units"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[7]/table/tbody/tr[5]/td[2]').text
                        print('10')
                        dict["Median_Home_Value"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[7]/table/tbody/tr[6]/td[2]').text
                        print('11')
                        dict["Total_Hoouseholds"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[7]/table/tbody/tr[11]/td[2]').text
                        print('12')
                        dict["Avarage_Households_Size"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[7]/table/tbody/tr[12]/td[2]').text
                        print('13')
                        dict["Family_Households"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[3]/div[1]/div[7]/table/tbody/tr[13]/td[2]').text
                        print('14')
                        print('HTML page params was copied to dict success------------------------------')
                    except:
                        try:
                            driver = self.driver
                            print('trying to locate elements with another third path 5')
                            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="bodycontainer"]/div[4]/div[2]/div[8]/table/tbody/tr[2]/td[2]')))
                            print('printing for checking the first element')
                            print(driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[2]/div[8]/table/tbody/tr[2]/td[2]').text)

                            time.sleep(2)
                            print('1')
                            dict["Total_Population"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[2]/div[8]/table/tbody/tr[2]/td[2]').text
                            print('2')
                            dict["Population_Growth_2010_2019"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[2]/div[12]/table/tbody/tr[2]/td[2]').text
                            print('3')
                            dict["Population_Growth_2019_2024"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[2]/div[12]/table/tbody/tr[2]/td[3]').text
                            print('4')
                            dict["Median_Household_Income"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[2]/div[8]/table/tbody/tr[11]/td[2]').text
                            print('5')
                            dict["Average_Household_Income"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[2]/div[8]/table/tbody/tr[12]/td[2]').text
                            print('6')
                            dict["Total_Housing_Units"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[2]/div[9]/table/tbody/tr[2]/td[2]').text
                            print('7')
                            dict["Owner_Occupied_HU"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[2]/div[9]/table/tbody/tr[3]/td[2]').text
                            print('8')
                            dict["Renter_Occupied_HU"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[2]/div[9]/table/tbody/tr[4]/td[2]').text
                            print('9')
                            dict["Vacant_Housing_Units"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[2]/div[9]/table/tbody/tr[5]/td[2]').text
                            print('10')
                            dict["Median_Home_Value"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[2]/div[9]/table/tbody/tr[6]/td[2]').text
                            print('11')
                            dict["Total_Hoouseholds"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[2]/div[9]/table/tbody/tr[11]/td[2]').text
                            print('12')
                            dict["Avarage_Households_Size"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[2]/div[9]/table/tbody/tr[12]/td[2]').text
                            print('13')
                            dict["Family_Households"] = driver.find_element_by_xpath('//*[@id="bodycontainer"]/div[4]/div[2]/div[9]/table/tbody/tr[13]/td[2]').text
                            print('14')
                            print('HTML page params was copied to dict success------------------------------')
                        except:
                            print('elements not found')
                            dict["Total_Population"] = "NA"
                            dict["Population_Growth_2010_2019"] = 'NA'
                            dict["Population_Growth_2019_2024"] = 'NA'
                            dict["Median_Household_Income"] = 'NA'
                            dict["Average_Household_Income"] = 'NA'
                            dict["Total_Housing_Units"] = 'NA'
                            dict["Owner_Occupied_HU"] = 'NA'
                            dict["Renter_Occupied_HU"] = 'NA'
                            dict["Vacant_Housing_Units"] = 'NA'
                            dict["Median_Home_Value"] = 'NA'
                            dict["Total_Hoouseholds"] = 'NA'
                            dict["Avarage_Households_Size"] = 'NA'
                            dict["Family_Households"] = 'NA'
        return dict

    def params_to_dict_block(self, dict):
        try:
            self.dict_block = dict.copy()
        except:
            print('Failed of HTML page params was copied to dict success')

    def params_to_dict_track(self, dict):
        try:
            self.dict_track = dict.copy()
        except:
            print('Failed of HTML page params was copied to dict success')

    def params_to_dict_zip_code(self, dict):
        try:
            self.dict_zip_code = dict.copy()
        except:
            print('Failed of HTML page params was copied to dict success')

    def params_to_dict_city(self, dict):
        try:
            self.dict_city = dict.copy()
        except:
            print('Failed of HTML page params was copied to dict success')

    def params_to_dict_county(self, dict):
        try:
            self.dict_county = dict.copy()
        except:
            print('Failed of HTML page params was copied to dict success')

    def params_to_dict_metro(self, dict):
        try:
            self.dict_metro = dict.copy()
        except:
            print('Failed of HTML page params was copied to dict success')

    # print all dictionaries
    def printall(self):
        print('hi all dictionaries\n')
        pp = pprint.PrettyPrinter(indent=4)
        print('basic info: {}'.format(self.dict_basic_info))
        pp.pprint(self.dict_basic_info)
        print('block: {}'.format(self.dict_block))
        pp.pprint(self.dict_block)
        print('track: {}'.format(self.dict_track))
        pp.pprint(self.dict_track)
        print('zip_code: {}'.format(self.dict_zip_code))
        pp.pprint(self.dict_zip_code)
        print('city: {}'.format(self.dict_city))
        pp.pprint(self.dict_city)
        print('metro: {}'.format(self.dict_metro))
        pp.pprint(self.dict_metro)

    # copy all dictionaries to xls file
    def xls_new_sheet_for_search_create(self):
        try:
            wb = openpyxl.load_workbook(self.xls_name)
            if wb.sheetnames.count(self.full_addr[:25]) == 0:
                example_sheet = wb["example"]
                wb.copy_worksheet(example_sheet)
                # print(wb.sheetnames)
                new_sheet = wb['example Copy']
                new_sheet.title = self.full_addr[:25]
                # print(wb.sheetnames)
                wb.save(self.xls_name)
                print("XLS new sheet is ready, sheet name: {}".format(self.full_addr[:25]))
                logging.debug("XLS new sheet is ready, sheet name: {}".format(self.full_addr[:25]))
                wb.close()
                return True
            else:
                print("address was already searched & exists in database")
                wb.close()
                logging.debug("address was already searched & exists in database")
                return False
        except:
            print('faild to create xls file')

    def basic_Info_dict_to_xls(self):
        try:
            # opening xls
            print('opening XLS to save params')
            logging.debug('opening XLS to save params')
            print('the xls file name is: {}'.format(self.xls_name))
            wb = openpyxl.load_workbook(self.xls_name)
            print(wb.sheetnames)
            sheet = wb[self.full_addr[:25]]
            print(self.full_addr[:25])
            # google link
            sheet['B17'].value = self.google_maps_link
            sheet['B3'].value = self.dict_basic_info['zip_code']
            wb.save(self.xls_name)
            wb.close()
            return True
        except:
            print('failed to copy info to XLS ')
            logging.debug('failed to open XLS')
            return False

    def all_dicts_to_xls(self):
        try:
            wb = openpyxl.load_workbook(self.xls_name)
            sheet = wb[self.full_addr[:25]]
            # print(wb.sheetnames)
            sheet['B7'].value = self.dict_block["Total_Population"]
            sheet['B8'].value = self.dict_block["Population_Growth_2010_2019"] + "(per year)"
            sheet['B9'].value = self.dict_block["Population_Growth_2019_2024"] + "(per year)"
            sheet['B10'].value = self.dict_block["Median_Household_Income"]
            sheet['B11'].value = self.dict_block["Average_Household_Income"]
            sheet['B12'].value = self.dict_block["Total_Housing_Units"]
            sheet['B13'].value = self.dict_block["Owner_Occupied_HU"]
            sheet['B14'].value = self.dict_block["Renter_Occupied_HU"]
            sheet['B15'].value = self.dict_block["Vacant_Housing_Units"]
            sheet['B16'].value = self.dict_block["Median_Home_Value"]
            sheet['B18'].value = self.dict_block["Total_Hoouseholds"]
            sheet['B19'].value = self.dict_block["Avarage_Households_Size"]
            sheet['B20'].value = self.dict_block["Family_Households"]

            sheet['C7'].value = self.dict_track["Total_Population"]
            sheet['C8'].value = self.dict_track["Population_Growth_2010_2019"] + "(per year)"
            sheet['C9'].value = self.dict_track["Population_Growth_2019_2024"] + "(per year)"
            sheet['C10'].value = self.dict_track["Median_Household_Income"]
            sheet['C11'].value = self.dict_track["Average_Household_Income"]
            sheet['C12'].value = self.dict_track["Total_Housing_Units"]
            sheet['C13'].value = self.dict_track["Owner_Occupied_HU"]
            sheet['C14'].value = self.dict_track["Renter_Occupied_HU"]
            sheet['C15'].value = self.dict_track["Vacant_Housing_Units"]
            sheet['C16'].value = self.dict_track["Median_Home_Value"]
            sheet['C18'].value = self.dict_track["Total_Hoouseholds"]
            sheet['C19'].value = self.dict_track["Avarage_Households_Size"]
            sheet['C20'].value = self.dict_track["Family_Households"]

            sheet['D7'].value = self.dict_zip_code["Total_Population"]
            sheet['D8'].value = self.dict_zip_code["Population_Growth_2010_2019"] + "(per year)"
            sheet['D9'].value = self.dict_zip_code["Population_Growth_2019_2024"] + "(per year)"
            sheet['D10'].value = self.dict_zip_code["Median_Household_Income"]
            sheet['D11'].value = self.dict_zip_code["Average_Household_Income"]
            sheet['D12'].value = self.dict_zip_code["Total_Housing_Units"]
            sheet['D13'].value = self.dict_zip_code["Owner_Occupied_HU"]
            sheet['D14'].value = self.dict_zip_code["Renter_Occupied_HU"]
            sheet['D15'].value = self.dict_zip_code["Vacant_Housing_Units"]
            sheet['D16'].value = self.dict_zip_code["Median_Home_Value"]
            sheet['D18'].value = self.dict_zip_code["Total_Hoouseholds"]
            sheet['D19'].value = self.dict_zip_code["Avarage_Households_Size"]
            sheet['D20'].value = self.dict_zip_code["Family_Households"]

            sheet['E7'].value = self.dict_city["Total_Population"]
            sheet['E8'].value = self.dict_city["Population_Growth_2010_2019"] + "(per year)"
            sheet['E9'].value = self.dict_city["Population_Growth_2019_2024"] + "(per year)"
            sheet['E10'].value = self.dict_city["Median_Household_Income"]
            sheet['E11'].value = self.dict_city["Average_Household_Income"]
            sheet['E12'].value = self.dict_city["Total_Housing_Units"]
            sheet['E13'].value = self.dict_city["Owner_Occupied_HU"]
            sheet['E14'].value = self.dict_city["Renter_Occupied_HU"]
            sheet['E15'].value = self.dict_city["Vacant_Housing_Units"]
            sheet['E16'].value = self.dict_city["Median_Home_Value"]
            sheet['E18'].value = self.dict_city["Total_Hoouseholds"]
            sheet['E19'].value = self.dict_city["Avarage_Households_Size"]
            sheet['E20'].value = self.dict_city["Family_Households"]

            sheet['F7'].value = self.dict_county["Total_Population"]
            sheet['F8'].value = self.dict_county["Population_Growth_2010_2019"] + "(per year)"
            sheet['F9'].value = self.dict_county["Population_Growth_2019_2024"] + "(per year)"
            sheet['F10'].value = self.dict_county["Median_Household_Income"]
            sheet['F11'].value = self.dict_county["Average_Household_Income"]
            sheet['F12'].value = self.dict_county["Total_Housing_Units"]
            sheet['F13'].value = self.dict_county["Owner_Occupied_HU"]
            sheet['F14'].value = self.dict_county["Renter_Occupied_HU"]
            sheet['F15'].value = self.dict_county["Vacant_Housing_Units"]
            sheet['F16'].value = self.dict_county["Median_Home_Value"]
            sheet['F18'].value = self.dict_county["Total_Hoouseholds"]
            sheet['F19'].value = self.dict_county["Avarage_Households_Size"]
            sheet['F20'].value = self.dict_county["Family_Households"]

            sheet['G7'].value = self.dict_metro["Total_Population"]
            sheet['G8'].value = self.dict_metro["Population_Growth_2010_2019"] + "(per year)"
            sheet['G9'].value = self.dict_metro["Population_Growth_2019_2024"] + "(per year)"
            sheet['G10'].value = self.dict_metro["Median_Household_Income"]
            sheet['G11'].value = self.dict_metro["Average_Household_Income"]
            sheet['G12'].value = self.dict_metro["Total_Housing_Units"]
            sheet['G13'].value = self.dict_metro["Owner_Occupied_HU"]
            sheet['G14'].value = self.dict_metro["Renter_Occupied_HU"]
            sheet['G15'].value = self.dict_metro["Vacant_Housing_Units"]
            sheet['G16'].value = self.dict_metro["Median_Home_Value"]
            sheet['G18'].value = self.dict_county["Total_Hoouseholds"]
            sheet['G19'].value = self.dict_county["Avarage_Households_Size"]
            sheet['G20'].value = self.dict_county["Family_Households"]

            wb.save(self.xls_name)
            wb.close()
            # printing the process
            print("Dictionaries was completed & saved in {}".format(self.xls_name))
            logging.debug("Dictionaries was completed & saved in {}".format(self.xls_name))
            return True
        except:
            print('Failed to copy to xls ')
            return False

    def return_dict_block(self):
        return self.dict_block

    def return_dict_basic_info(self):
        return self.dict_basic_info

    def return_dict_track(self):
        return self.dict_track

    def return_dict_zip_code(self):
        return self.dict_zip_code

    def return_dict_city(self):
        return self.dict_city

    def return_dict_county(self):
        return self.dict_county

    def return_dict_metro(self):
        return self.dict_metro

    def return_county_name(self):
        return self.county

    def return_zip_code_for_zillow_use(self):
        return self.dict_basic_info['zip_code']



