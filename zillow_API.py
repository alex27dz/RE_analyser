'''
CMA Class was build for adding the CMA / Zillow information of the given address from input
Using address to collect information from Zillow API and store it in dictionary
'''
'''
import zillow
import locale
import json
import openpyxl
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import pprint
import logging
from selenium.webdriver.support.ui import Select

# my api_key = "X1-ZWz1hbswvtw74b_3tnpx"


class CMA(object):
    def __init__(self, address, api_key, zip_code, xls_name):
        # all setup params
        self.address = address
        self.api_key = api_key
        self.zip_code = zip_code
        self.xls_name = xls_name

        self.dict_zillow = {
            'address': self.address,
            'api_key': self.api_key,
            'zip_code': self.zip_code,
            'zpid': '',
            'link - comparables': '',
            'link - graphs_and_data': '',
            'link - home_details': '',
            'link - map_this_home': '',
            'amount': '',
            'amount_currency': '',
            'amount_last_updated': '',
            'valuation_range_high': '',
            'valuation_range_low': '',
            'bathrooms': '',
            'bedrooms': '',
            'complete': '',
            'finished_sqft': '',
            'fips_county': '',
            'last_sold_date': '',
            'last_sold_price': '',
            'lot_size_sqft': '',
            'tax_assessment': '',
            'tax_assessment_year': '',
            'usecode': '',
            'year_built': '',

        }
    # activation zillow API and copy params to dict
    def zillow_api(self):
        try:
            locale.setlocale(locale.LC_ALL, '')
            api = zillow.ValuationApi()
            # get deep search results, also getting the zswid-ID
            data = api.GetDeepSearchResults(self.api_key, self.address, self.zip_code)
            my_dict = data.get_dict()
            # copy web page params to dictionary dict_zillow
            self.dict_zillow['zpid'] = my_dict['zpid']
            self.dict_zillow['link - comparables'] = my_dict['links']['comparables']
            self.dict_zillow['link - graphs_and_data'] = my_dict['links']['graphs_and_data']
            self.dict_zillow['link - home_details'] = my_dict['links']['home_details']
            self.dict_zillow['link - map_this_home'] = my_dict['links']['map_this_home']
            self.dict_zillow['amount'] = my_dict['zestimate']['amount']
            self.dict_zillow['amount_currency'] = my_dict['zestimate']['amount_currency']
            self.dict_zillow['amount_last_updated'] = my_dict['zestimate']['amount_last_updated']
            self.dict_zillow['valuation_range_high'] = my_dict['zestimate']['valuation_range_high']
            self.dict_zillow['valuation_range_low'] = my_dict['zestimate']['valuation_range_low']
            self.dict_zillow['bathrooms'] = my_dict['extended_data']['bathrooms']
            self.dict_zillow['bedrooms'] = my_dict['extended_data']['bedrooms']
            self.dict_zillow['complete'] = my_dict['extended_data']['complete']
            self.dict_zillow['finished_sqft'] = my_dict['extended_data']['finished_sqft']
            self.dict_zillow['fips_county'] = my_dict['extended_data']['fips_county']
            self.dict_zillow['last_sold_date'] = my_dict['extended_data']['last_sold_date']
            self.dict_zillow['last_sold_price'] = my_dict['extended_data']['last_sold_price']
            self.dict_zillow['lot_size_sqft'] = my_dict['extended_data']['lot_size_sqft']
            self.dict_zillow['tax_assessment_year'] = my_dict['extended_data']['tax_assessment_year']
            self.dict_zillow['usecode'] = my_dict['extended_data']['usecode']
            self.dict_zillow['year_built'] = my_dict['extended_data']['year_built']
            return True
        except:
            print('fail to get params from zillow api')
            logging.debug('fail')
            self.dict_zillow['zpid'] = 'fail to get params from zillow api'
            return False

    # printing all dicts
    def print_all(self):
        pp = pprint.PrettyPrinter(indent=4)
        pp.pprint(self.dict_zillow)
    # returning all dictionaries for future use for general list
    def return_dict_zillow(self):
        return self.dict_zillow
    # copy dict to xls file
    def xls_new_sheet_for_search_create(self):
        wb = openpyxl.load_workbook(self.xls_name)
        if wb.sheetnames.count(self.address[:25]) == 0:
            example_sheet = wb["example"]
            wb.copy_worksheet(example_sheet)
            # print(wb.sheetnames)
            new_sheet = wb['example Copy']
            new_sheet.title = self.address[:25]
            # print(wb.sheetnames)
            wb.save(self.xls_name)
            print("XLS new sheet is ready, sheet name: {}".format(self.address[:25]))
            logging.debug("XLS new sheet is ready, sheet name: {}".format(self.address[:25]))
            wb.close()
            return True
        else:
            print("address was already searched & exists in database")
            logging.debug("address was already searched & exists in database")
            return False
    def all_dicts_to_xls(self):
        try:
            wb = openpyxl.load_workbook(self.xls_name)
            sheet = wb[self.address[:25]]
            # print(wb.sheetnames)
            sheet['B2'].value = self.dict_zillow['address']
            sheet['B3'].value = self.dict_zillow['api_key']
            sheet['B4'].value = self.dict_zillow['zip_code']
            sheet['B5'].value = self.dict_zillow['zpid']
            sheet['B7'].value = self.dict_zillow['link - comparables']
            sheet['B8'].value = self.dict_zillow['link - graphs_and_data']
            sheet['B9'].value = self.dict_zillow['link - home_details']
            sheet['B10'].value = self.dict_zillow['link - map_this_home']
            sheet['B12'].value = self.dict_zillow['amount']
            sheet['B13'].value = self.dict_zillow['amount_currency']
            sheet['B14'].value = self.dict_zillow['amount_last_updated']
            sheet['B15'].value = self.dict_zillow['valuation_range_high']
            sheet['B16'].value = self.dict_zillow['valuation_range_low']
            sheet['B18'].value = self.dict_zillow['bathrooms']
            sheet['B19'].value = self.dict_zillow['bedrooms']
            sheet['B20'].value = self.dict_zillow['complete']
            sheet['B21'].value = self.dict_zillow['finished_sqft']
            sheet['B22'].value = self.dict_zillow['fips_county']
            sheet['B23'].value = self.dict_zillow['last_sold_date']
            sheet['B24'].value = self.dict_zillow['last_sold_price']
            sheet['B25'].value = self.dict_zillow['lot_size_sqft']
            sheet['B26'].value = self.dict_zillow['tax_assessment']
            sheet['B27'].value = self.dict_zillow['tax_assessment_year']
            sheet['B28'].value = self.dict_zillow['usecode']
            sheet['B29'].value = self.dict_zillow['year_built']
            wb.save(self.xls_name)
            wb.close()
            # printing the process
            print("Dictionaries was completed & saved in {}".format(self.xls_name))
            logging.debug("Dictionaries was completed & saved in {}".format(self.xls_name))
            return True
        except:
            return False






'''





























